"""
週毎SISデータ一覧（2014-2026）から、機種一覧_MY_コイン単価.xlsx の
SIS平均アウト / SIS割数% を空欄行に補完する（SIS平均MYは計算式不明のためスキップ）。

実行: python fill_sis_avg.py
"""
import os
import sys
import io
import openpyxl
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

MASTER = r"C:\Users\h.kadoya\Desktop\slocri\ai収集\機種一覧_MY_コイン単価.xlsx"

# 名称揺れの手動マッピング (master名 → SIS名)
ALIAS = {
    "絶対衝激4": "L絶対衝激IV",
    "エヴァンゲリオンー約束の扉ー": "LBエヴァンゲリヲン約束の扉",
    "ネオアイムジャグラー": "SネオアイムジャグラーEX",
    "マタドール|||": "LBマタドールIII",
    "咲～頂上決戦～": "L咲-Saki-頂上決戦",
}
SIS_DIRS = [
    "Z:/01_SISデータ/PS",
    "Z:/01_SISデータ/PS/過去データまとめ",
]


def collect_sis_weekly_files():
    files = []
    for d in SIS_DIRS:
        if not os.path.isdir(d):
            continue
        for f in os.listdir(d):
            if f.endswith(".xlsm") and "週毎" in f and not f.startswith("~$"):
                files.append(os.path.join(d, f))
    return sorted(files)


def normalize(s: str) -> str:
    """機種名比較用の正規化"""
    s = s.strip()
    # 全角/半角ノイズ除去
    for ch in [" ", "　", "／", "/", "-", "ー", "～", "~"]:
        s = s.replace(ch, "")
    return s.lower()


def keys_for(s: str):
    """機種名から (strict_key, noprefix_key, has_prefix) を返す"""
    n = normalize(s)
    no_pfx = n
    has_pfx = False
    for pfx in ("lb", "sb", "l", "s"):
        if n.startswith(pfx):
            no_pfx = n[len(pfx):]
            has_pfx = True
            break
    return n, no_pfx, has_pfx


def collect_machine_weekly_data(sis_files, targets):
    """targets = [(row, master_name, strict, noprefix, has_pfx), ...]
    マッチング:
      - strict一致 → 採用
      - どちらか一方にプレフィックスがない場合のみ noprefix一致を採用
    """
    # strict_lookup[strict_key] = (row, master_name)
    strict_lookup = {t[2]: (t[0], t[1]) for t in targets}
    # noprefix_lookup[noprefix_key] = list of (row, master_name, has_pfx)
    noprefix_lookup = defaultdict(list)
    for t in targets:
        noprefix_lookup[t[3]].append((t[0], t[1], t[4]))

    data = defaultdict(list)  # row → [(out, rate, sis_name)]
    raw_match = {}  # row → actual SIS machine name

    for path in sis_files:
        print(f"  読込: {os.path.basename(path)}")
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=True)
        except Exception as e:
            print(f"   skip: {e}")
            continue

        week_sheets = [s for s in wb.sheetnames if "~" in s]
        for sname in week_sheets:
            ws = wb[sname]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 12:
                    continue
                machine = row[1]
                if not machine or not isinstance(machine, str):
                    continue
                out = row[6]
                rate = row[11]
                if not isinstance(out, (int, float)) or not isinstance(rate, (int, float)):
                    continue

                sis_strict, sis_noprefix, sis_has_pfx = keys_for(machine)

                # 1) strict一致
                if sis_strict in strict_lookup:
                    target_row, _ = strict_lookup[sis_strict]
                    data[target_row].append((out, rate, machine))
                    raw_match.setdefault(target_row, machine)
                    continue

                # 2) noprefix一致（少なくとも一方がプレフィックス無し）
                for target_row, _, master_has_pfx in noprefix_lookup.get(sis_noprefix, []):
                    if not (master_has_pfx and sis_has_pfx):
                        # どちらか or 両方プレフィックス無 → noprefixマッチOK
                        data[target_row].append((out, rate, machine))
                        raw_match.setdefault(target_row, machine)
                        break
        wb.close()
    return data, raw_match


def main():
    wb = openpyxl.load_workbook(MASTER)
    ws = wb.active

    # SIS空欄行（SIS平均アウト=col12, SIS平均MY=col13, SIS割数%=col14 すべて空）を抽出
    empty_rows = []
    for r in range(3, ws.max_row + 1):
        machine = ws.cell(row=r, column=1).value
        out_v = ws.cell(row=r, column=12).value
        my_v = ws.cell(row=r, column=13).value
        ratio_v = ws.cell(row=r, column=14).value
        if not machine:
            continue
        if out_v is None and my_v is None and ratio_v is None:
            empty_rows.append((r, str(machine).strip()))

    print(f"SIS空欄: {len(empty_rows)}行")

    targets = []
    for r, m in empty_rows:
        match_name = ALIAS.get(m, m)
        strict, noprefix, has_pfx = keys_for(match_name)
        targets.append((r, m, strict, noprefix, has_pfx))

    sis_files = collect_sis_weekly_files()
    print(f"週毎SISファイル: {len(sis_files)}件")

    data, raw_match = collect_machine_weekly_data(sis_files, targets)

    updated = 0
    not_found = []
    for r, master_name, _, _, _ in targets:
        weeks = data.get(r, [])
        if not weeks:
            not_found.append(master_name)
            continue
        outs = [w[0] for w in weeks]
        rates = [w[1] for w in weeks]
        avg_out = round(sum(outs) / len(outs))
        avg_rate = round(sum(rates) / len(rates), 1)

        ws.cell(row=r, column=12).value = avg_out
        ws.cell(row=r, column=14).value = avg_rate
        updated += 1
        print(f"  ✓ {master_name} ← (SIS: {raw_match[r]}) アウト={avg_out}, 割数={avg_rate}% [{len(weeks)}週]")

    wb.save(MASTER)
    print(f"\n更新完了: {updated}/{len(empty_rows)}件")
    print(f"未マッチ: {len(not_found)}件")
    for m in not_found[:30]:
        print(f"  - {m}")
    if len(not_found) > 30:
        print(f"  ... 他{len(not_found)-30}件")


if __name__ == "__main__":
    main()
