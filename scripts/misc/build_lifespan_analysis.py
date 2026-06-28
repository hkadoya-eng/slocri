# -*- coding: utf-8 -*-
"""寿命別 初月挙動の比較: ダメ台 vs 優秀台
週次SIS全件(_wk_p*.json)から、各機種の first/last/総週数/持続率/台数推移/割数を算出。
成熟機種(初週<=2025-09=もう生死が出ている)を寿命でバケツ分けして、初月の死相を探る。
"""
import json, io, glob, statistics
from collections import defaultdict

rows=[]
for fp in sorted(glob.glob("_wk_p*.json")):
    try:
        with io.open(fp,encoding="utf-8") as f: rows+=json.load(f)
    except Exception as e: print("skip",fp,e)
print("総行数",len(rows))

def f(x):
    try: return float(x)
    except: return None

M=defaultdict(list)
for r in rows: M[r["machine"]].append(r)
for m in M: M[m].sort(key=lambda x:x["week_start"])

recs=[]
for m,rs in M.items():
    def out(i): return f(rs[i]["out_coins"]) if len(rs)>i else None
    def mc(i):  return f(rs[i]["avg_machine_count"]) if len(rs)>i else None
    w1,w4,w8=out(0),out(3),out(7)
    c1,c4,c8=mc(0),mc(3),mc(7)
    clast=mc(len(rs)-1)
    pr=[f(r["payout_rate"]) for r in rs[:8] if f(r["payout_rate"]) is not None]
    recs.append({
        "機種":m,"初週":rs[0]["week_start"],"最終":rs[-1]["week_start"],"総週数":len(rs),
        "現役": rs[-1]["week_start"]>="2026-06-01",
        "初週out": round(w1) if w1 else None,
        "持続率4": round(w4/w1*100) if (w1 and w4) else None,
        "持続率8": round(w8/w1*100) if (w1 and w8) else None,
        "台数初週": c1,
        "台数4週%": round((c4/c1-1)*100) if (c1 and c4) else None,
        "台数8週%": round((c8/c1-1)*100) if (c1 and c8) else None,
        "割数初8": round(statistics.mean(pr),1) if pr else None,
    })

mat=[r for r in recs if r["初週"]<="2025-09-01"]
def bkt(n): return "短命≤13w(ダメ)" if n<=13 else ("長寿≥45w(優秀)" if n>=45 else "中27前後")
G=defaultdict(list)
for r in mat: G[bkt(r["総週数"])].append(r)

def stat(items,k):
    v=[i[k] for i in items if i.get(k) is not None]
    if not v: return (None,None)
    return (round(statistics.mean(v),1), round(statistics.median(v),1))

print(f"\n成熟機種(初週<=2025-09)= {len(mat)} 機種")
print(f'{"バケツ":<16}{"n":>4}{"持続率4(平均/中)":>16}{"持続率8":>10}{"台数4週%":>10}{"台数8週%":>10}{"割数":>8}')
for b in ["短命≤13w(ダメ)","中27前後","長寿≥45w(優秀)"]:
    g=G[b]
    r4=stat(g,"持続率4"); r8=stat(g,"持続率8"); c4=stat(g,"台数4週%"); c8=stat(g,"台数8週%"); wr=stat(g,"割数初8")
    print(f'{b:<16}{len(g):>4}{str(r4[0])+"/"+str(r4[1]):>16}{str(r8[0]):>10}{str(c4[0]):>10}{str(c8[0]):>10}{str(wr[0]):>8}')

print("\n=== 最短命ダメ台(成熟) ===")
for r in sorted([x for x in mat if x["総週数"]<=13],key=lambda x:x["総週数"])[:10]:
    print(f' {r["総週数"]:>2}週 持続4={r["持続率4"]} 台数4週={r["台数4週%"]}% 割数={r["割数初8"]} {r["機種"][:26]}')

print("\n=== 参照: 北斗/モンキー ===")
for r in mat:
    if "北斗の拳" in r["機種"] and "転生" not in r["機種"] and "無双" not in r["機種"] or "モンキーターン" in r["機種"]:
        print(f' {r["総週数"]:>3}週 現役={r["現役"]} 持続4={r["持続率4"]} 持続8={r["持続率8"]} 台数4={r["台数4週%"]}% 台数8={r["台数8週%"]}% {r["機種"][:24]}')

# Excel
cols=["機種","初週","最終","総週数","現役","初週out","持続率4","持続率8","台数初週","台数4週%","台数8週%","割数初8"]
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font,PatternFill,Alignment
    from openpyxl.utils import get_column_letter
    wb=Workbook(); ws=wb.active; ws.title="寿命×持続率×台数"
    ws.append(["成熟機種(初週<=2025-09)。持続率n=週n out/初週out。台数%=初週比。週次SISより自動算出。"])
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(cols))
    ws.cell(1,1).font=Font(italic=True,color="996600")
    ws.append(cols); hr=ws.max_row
    for c in range(1,len(cols)+1):
        cell=ws.cell(hr,c); cell.font=Font(bold=True,color="FFFFFF")
        cell.fill=PatternFill("solid",fgColor="444444"); cell.alignment=Alignment(horizontal="center")
    for r in sorted(mat,key=lambda x:x["総週数"],reverse=True):
        ws.append([r[c] for c in cols]); rr=ws.max_row
        nm=r["機種"]
        if ("北斗の拳" in nm and "転生" not in nm and "無双" not in nm) or "モンキーターン" in nm:
            for c in range(1,len(cols)+1): ws.cell(rr,c).fill=PatternFill("solid",fgColor="FFF2CC")
        elif r["総週数"]<=13:
            for c in range(1,len(cols)+1): ws.cell(rr,c).fill=PatternFill("solid",fgColor="FCE4E4")
    for i,w in enumerate([30,11,11,7,7,9,8,8,9,9,9,8],1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="A3"
    wb.save("ai収集/分析_寿命_持続率_台数_v0.2.xlsx")
    print("\n出力: ai収集/分析_寿命_持続率_台数_v0.2.xlsx")
except ImportError:
    print("openpyxl無し: Excel出力スキップ")
