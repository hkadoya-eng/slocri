# -*- coding: utf-8 -*-
"""新ゲーム性提案「自力ブースト×やめ時消去」型 + 根拠をExcel化。"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb=Workbook()
H=PatternFill("solid",fgColor="444444"); HF=Font(bold=True,color="FFFFFF")
def hdr(ws,row,cols):
    for i,c in enumerate(cols,1):
        x=ws.cell(row,i,c); x.font=HF; x.fill=H; x.alignment=Alignment(horizontal="center")

# ---- 提案シート ----
ws=wb.active; ws.title="提案"
L=[
 ("新ゲーム性提案『自力ブースト×やめ時消去』型",True,"D85A30",14),
 ("（スロキー編集部発・オリジナル試算／実機に数値を紐付けない）",False,"888888",10),
 ("",False,"",11),
 ("【設計思想】スペック起点でなく“リテンション起点”。台の生死を決めるのは機械割/初当り/一撃ではなく『持続率×稼働値＝打ち続けられるか』（今回の稼働データ分析の結論）。",False,"333333",11),
]
r=1
for t,b,c,sz in L:
    cell=ws.cell(r,1,t); cell.font=Font(bold=b,color=c or "000000",size=sz); cell.alignment=Alignment(wrap_text=True,vertical="top"); r+=1
ws.column_dimensions["A"].width=30
for col in "BCDE": ws.column_dimensions[col].width=22

r+=1
ws.cell(r,1,"■ 市場で勝ってる3要素＋1を合成").font=Font(bold=True,color="D85A30"); r+=1
hdr(ws,r,["要素","実例","効く理由"]); r+=1
for a,b,c in [
 ("自力介入","モンキーターン/虚構推理","打ち手が関与→離脱しない→稼働持続"),
 ("レア役ブースト状態","東京喰種(赫眼)","『今引けば強い』引き所の快感"),
 ("やめ時消去","虚構推理(強制ループ)","物理的にやめられない→持続率↑"),
 ("二層分散","モンキーターン","普段マイルド→ライト層が焼けない→持続"),
 ("【避ける】一撃ロマンの高アバレ単発","（多数の短命台）","初月で資金焼き切りライト層蒸発→短命"),
]:
    ws.cell(r,1,a); ws.cell(r,2,b); ws.cell(r,3,c); r+=1

r+=1
ws.cell(r,1,"■ 仕様の核").font=Font(bold=True,color="D85A30"); r+=1
for s in [
 "通常時＝レア役ブースト状態を搭載。押し順/狙い目の的中でブーストG数を“自分で延ばせる”（受け身の高確でなく打ち手が育てる高確）",
 "AT＝二段構え。メインはマイルドに刻み資金を焼かない／上位へは自力のチャンス目集中で入りにいける",
 "AT後＝必ずブースト高確＋引き継ぎCZに接続し“やめ時消去”",
 "設定看破要素あり＝ライト(介入で楽しい)と玄人(判別で楽しい)の両取り＝持続率の土台",
]:
    c=ws.cell(r,1,"・"+s); c.alignment=Alignment(wrap_text=True); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5); r+=1

r+=1
ws.cell(r,1,"■ 数値設計（オリジナル試算）").font=Font(bold=True,color="D85A30"); r+=1
hdr(ws,r,["区分","純増(枚/G)","1セットG","継続率","平均連(=1/(1-継続))","平均出玉(=純増×G×連)"]); r+=1
for a,b,c,d2,e,f in [
 ("メインAT",2.5,40,"65%","約2.86連","約286枚"),
 ("上位AT",5.0,40,"85%","約6.67連","約1333枚"),
]:
    for i,v in enumerate([a,b,c,d2,e,f],1): ws.cell(r,i,v)
    r+=1
for s in ["初当り 約1/290(設6)／コイン単価 約3.2円／天井 700G+周期","機械割 設定1 約97.5%〜設定6 約114.5%（現実レンジ97〜114%内）／上位85%ループでコンプリ(19000枚)も射程"]:
    c=ws.cell(r,1,s); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); r+=1

r+=1
ws.cell(r,1,"■ 自己検算").font=Font(bold=True,color="1f9d4d"); r+=1
for s in ["平均連チャン≒1/(1-継続率): メイン1/(1-0.65)=2.86 ✓ 上位1/(1-0.85)=6.67 ✓","平均出玉≒純増×1セットG×平均連: メイン2.5×40×2.86≒286 ✓ 上位5×40×6.67≒1333 ✓","機械割は現実レンジ97〜114%内に収束 ✓"]:
    c=ws.cell(r,1,s); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); r+=1

r+=1
ws.cell(r,1,"■ 正直な穴").font=Font(bold=True,color="C77B00"); r+=1
for s in ["“流行りそう”はデータ起点の仮説。良い仕様でも台数を盛りすぎれば死ぬ(供給規律)・型式試験の運もある","最終判定は打ち手の審美眼＝『実際に打って痺れるか』。構造で持続は言えても体感は打った人にしか分からない"]:
    c=ws.cell(r,1,"・"+s); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); r+=1

# ---- 根拠シート ----
ws2=wb.create_sheet("根拠(今回の分析)")
ws2.column_dimensions["A"].width=30; ws2.column_dimensions["B"].width=70
ws2.cell(1,1,"提案の根拠＝2026-06〜07の稼働データ分析で判明した事実").font=Font(bold=True,color="D85A30")
hdr(ws2,2,["論点","内容"])
rr=3
for a,b in [
 ("生死はスペック無関係","機械割/MY/初当りは寿命とほぼ無相関(MY r=-0.03・初当り同値で運命真逆)。決めるのは持続率×稼働値"),
 ("持続率が最強シグナル","初月持続率(4週÷初週アウト) 長寿平均66% vs 短命47%。2週持続でも長寿86 vs 短命72"),
 ("稼働値(人気)が第2軸","アウト÷週中央値。長寿335% vs 短命262%。持続率と別軸で効く"),
 ("長寿は希少","公式稼働貢献週で長寿≥45週は8機種のみ・短命≤13週が102機種(大多数)"),
 ("死に台は撤去されず残る","真北斗無双=公式貢献5週なのに設置100週。設置週数は死を隠す→稼働貢献週(公式)で見る"),
 ("レア役ブースト状態型が旬","東京喰種赫眼=下段リプレイで10〜50G・レア役出現率が約1/2.6へ。化物語/攻殻も同系"),
 ("2週で予測可能","2週判定[強]=最終半年超率25%(中央14週) vs [中弱]=2%。強=稼働値≥260かつ持続率≥83"),
 ("スペック値の罠","機械割114.9%は規則の表示上限で全台横並び=比較情報ゼロ。型式試験は181万/回・適合率10〜20%"),
]:
    ws2.cell(rr,1,a); c=ws2.cell(rr,2,b); c.alignment=Alignment(wrap_text=True); rr+=1
ws2.freeze_panes="A3"

wb.save("ai収集/分析_新ゲーム性提案_自力ブースト×やめ時消去_v1.xlsx")
print("出力: ai収集/分析_新ゲーム性提案_自力ブースト×やめ時消去_v1.xlsx")
