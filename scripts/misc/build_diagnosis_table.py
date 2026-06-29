# -*- coding: utf-8 -*-
"""新台診断テーブル v0.5（稼働貢献週を SIS公式値=sis_machine_stats.contrib_weeks に修正）
重大修正: 旧版は週次行数(len)を稼働貢献週として誤用していた(死に台も設置週として水増し)。
正: 稼働貢献週=公式contrib_weeks。設置週数(週次行数)は参考として別掲。持続率/稼働値は週次outから(従来通り有効)。
入力: _wk_p*.json + _stats.json / 出力: ai収集/分析_新台診断_v0.3.xlsx"""
import json, io, glob, statistics
from collections import defaultdict

rows=[]
for fp in sorted(glob.glob("_wk_p*.json")):
    with io.open(fp,encoding="utf-8") as f: rows+=json.load(f)
if not rows:
    print("週次ファイル無し"); raise SystemExit

# 公式 稼働貢献週(機種名は同一インポート元なので完全一致で結合)
stats={}
try:
    with io.open("_stats.json",encoding="utf-8") as f:
        for r in json.load(f):
            if r.get("machine") and r["machine"]!="__config__":
                stats[r["machine"]]=r.get("contrib_weeks")
except FileNotFoundError:
    print("警告: _stats.json なし → 稼働貢献週が出せない");

def f_(x):
    try: return float(x)
    except: return None

wk_med={}
_t=defaultdict(list)
for r in rows:
    o=f_(r["out_coins"])
    if o: _t[r["week_start"]].append(o)
for w,a in _t.items(): s=sorted(a); wk_med[w]=s[len(s)//2]

M=defaultdict(list)
for r in rows: M[r["machine"]].append(r)
for m in M: M[m].sort(key=lambda x:x["week_start"])

def jittai(ret4):
    if ret4 is None: return "計測中"
    if ret4>=66: return "健全"
    if ret4>=50: return "微妙"
    return "実質死亡"

R2_OK,R2_WARN=83,73
def grade(ret4,ret2,k):
    if ret4 is not None:
        return "優秀" if ret4>=66 else ("注意" if ret4>=50 else "危険")
    if ret2 is not None:
        pOK,pBad=ret2>=R2_OK,ret2<R2_WARN
        dOK=(k is None or k>=260); dBad=(k is not None and k<190)
        if pOK and dOK: return "優秀(暫定)"
        if pBad or dBad: return "危険(暫定)"
        return "注意(暫定)"
    return "計測中"

recs=[]
for m,rs in M.items():
    def out(i): return f_(rs[i]["out_coins"]) if len(rs)>i else None
    def mc(i):  return f_(rs[i]["avg_machine_count"]) if len(rs)>i else None
    w1,w2,w4,w8=out(0),out(1),out(3),out(7)
    c1,cL=mc(0),mc(len(rs)-1)
    peakv=[f_(r["avg_machine_count"]) for r in rs if f_(r["avg_machine_count"]) is not None]
    b1,b2=wk_med.get(rs[0]["week_start"]),(wk_med.get(rs[1]["week_start"]) if len(rs)>1 else None)
    ret2=round(w2/w1*100) if (w1 and w2) else None
    ret4=round(w4/w1*100) if (w1 and w4) else None
    ret8=round(w8/w1*100) if (w1 and w8) else None
    kat1=round(w1/b1*100) if (w1 and b1) else None
    kat2=round(w2/b2*100) if (w2 and b2) else None
    k=kat2 if kat2 is not None else kat1
    contrib=stats.get(m)              # 公式 稼働貢献週
    setti=len(rs)                     # 設置週数(週次行数・参考)
    recs.append({
        "機種":m,"初週":rs[0]["week_start"],
        "稼働貢献週(公式)":contrib,"設置週数":setti,
        "死に台期間": (setti-contrib) if (contrib is not None) else None,  # 設置-貢献=惰性設置
        "初月実態":jittai(ret4),
        "初週稼働値":kat1,"2週稼働値":kat2,
        "2週持続率":ret2,"初月持続率(4週)":ret4,"8週持続率":ret8,
        "台数初週":round(c1,1) if c1 else None,"台数現在":round(cL,1) if cL else None,
        "台数ピーク":round(max(peakv),1) if peakv else None,
        "台数推移%":round((cL/c1-1)*100) if (c1 and cL) else None,
        "判定":grade(ret4,ret2,k),
    })

# ===== 公式稼働貢献週での再検証(持続率は効くか) =====
mat=[r for r in recs if r["初週"]<="2025-09-01" and r["稼働貢献週(公式)"] is not None]
def bkt(n): return "短命<=13" if n<=13 else ("長寿>=45" if n>=45 else "中")
G=defaultdict(list)
for r in mat: G[bkt(r["稼働貢献週(公式)"])].append(r)
def gm(items,k2):
    v=[i[k2] for i in items if i.get(k2) is not None]
    return round(statistics.mean(v),1) if v else None
print("=== 【公式稼働貢献週】での寿命別 持続率/稼働値（再検証） ===")
for b in ["短命<=13","中","長寿>=45"]:
    g=G[b]; print(f'  {b}: n={len(g):>3} 初週稼働値={gm(g,"初週稼働値")} 2週持続={gm(g,"2週持続率")} 4週持続={gm(g,"初月持続率(4週)")}')
def corr(xs,ys):
    n=len(xs)
    if n<3: return None
    mx,my=statistics.mean(xs),statistics.mean(ys)
    cov=sum((x-mx)*(y-my) for x,y in zip(xs,ys))/n
    sx,sy=statistics.pstdev(xs),statistics.pstdev(ys)
    return round(cov/((sx*sy) or 1),2)
p4=[(r["初月持続率(4週)"],r["稼働貢献週(公式)"]) for r in mat if r["初月持続率(4週)"] is not None]
print(f'  相関 4週持続率→公式稼働貢献週: r={corr([a for a,_ in p4],[b for _,b in p4])}')

# 公式 vs 旧カウントの乖離トップ
print("\n=== 死に台期間(設置-公式貢献)が長い=旧版が水増ししてた台 ===")
for r in sorted([x for x in recs if x["死に台期間"] is not None],key=lambda x:-x["死に台期間"])[:8]:
    print(f'  公式{r["稼働貢献週(公式)"]}週 / 設置{r["設置週数"]}週 (惰性{r["死に台期間"]}週)  {r["機種"][:28]}')

# ===== Excel =====
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font,PatternFill,Alignment
    from openpyxl.utils import get_column_letter
    wb=Workbook()
    ws0=wb.active; ws0.title="説明"
    L=[
        ("新台診断テーブル v0.5（稼働貢献週=SIS公式値に修正）",True,"D85A30",14),
        ("",False,"",11),
        ("【重大修正 2026-06-29】旧版は週次データの行数を『稼働貢献週』として誤用していた。死に台でも設置されてる限り週次行が残るため大幅に水増しされていた(例:真北斗無双 誤100週→正5週)。",False,"D03030",11),
        ("",False,"",11),
        ("● 稼働貢献週(公式) = sis_machine_stats のSIS公式値。SISが『○週稼動貢献中』と認めた週数(稼働が一定水準を満たした週)。これが正。",True,"333333",11),
        ("● 設置週数 = 週次データに行がある週数(参考)。死に台でも設置され続ければ伸びる。設置週数≠生死。",False,"555555",11),
        ("● 死に台期間 = 設置週数 − 公式稼働貢献週。大きいほど『貢献しなくなった後も惰性で設置された』台。",False,"C77B00",11),
        ("",False,"",11),
        ("● 稼働値 = アウト ÷ その週の全機種中央値(人気/需要)。持続率 = ◯週目アウト ÷ 初週アウト(定着)。※週次outは『平均IN』=1台平均値で総量ではない。比率なので診断は有効。",True,"333333",11),
        ("● 初月実態 = 初月持続率による生死(健全≥66/微妙50-66/実質死亡<50/計測中)。",True,"333333",11),
        ("● 判定 = 4週揃えば持続率4週(優秀≥66/注意≥50/危険<50)、新台は2週暫定(稼働値≥260かつ持続率≥83で優秀)。",True,"333333",11),
        ("",False,"",11),
        ("【データ源】SIS週次(sis_weekly_data)+稼働貢献週(sis_machine_stats)。元はZ:/01_SISデータ/PS/週毎SISデータ一覧_2026.xlsm。生成: scripts/misc/build_diagnosis_table.py。",False,"888888",10),
    ]
    for i,(t,b,c,sz) in enumerate(L,1):
        cell=ws0.cell(i,1,t); cell.font=Font(bold=b,color=c or "000000",size=sz)
        cell.alignment=Alignment(wrap_text=True,vertical="top")
    ws0.column_dimensions["A"].width=118

    cols=["機種","初週","稼働貢献週(公式)","設置週数","死に台期間","初月実態","初週稼働値","2週稼働値",
          "2週持続率","初月持続率(4週)","8週持続率","台数初週","台数現在","台数ピーク","台数推移%","判定"]
    ws=wb.create_sheet("新台診断データ")
    ws.append(["※直近導入順。稼働貢献週=SIS公式値(これが正)。設置週数は参考。意味は『説明』シート参照。"])
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(cols))
    ws.cell(1,1).font=Font(italic=True,color="996600")
    ws.append(cols); hr=ws.max_row
    for c in range(1,len(cols)+1):
        cell=ws.cell(hr,c); cell.font=Font(bold=True,color="FFFFFF")
        cell.fill=PatternFill("solid",fgColor="444444"); cell.alignment=Alignment(horizontal="center")
    GC={"優秀":"E3F5E9","注意":"FFF3DC","危険":"FCE4E4","優秀(暫定)":"E3F5E9","注意(暫定)":"FFF3DC","危険(暫定)":"FCE4E4","計測中":"ECECEC"}
    JC={"健全":"E3F5E9","微妙":"FFF3DC","実質死亡":"FCE4E4","計測中":"ECECEC"}
    for r in sorted(recs,key=lambda x:x["初週"],reverse=True):
        ws.append([r[c] for c in cols]); rr=ws.max_row
        f=GC.get(r["判定"])
        if f: ws.cell(rr,cols.index("判定")+1).fill=PatternFill("solid",fgColor=f)
        jf=JC.get(r["初月実態"])
        if jf: ws.cell(rr,cols.index("初月実態")+1).fill=PatternFill("solid",fgColor=jf)
        if r["死に台期間"] is not None and r["死に台期間"]>=20:
            c0=ws.cell(rr,cols.index("死に台期間")+1); c0.fill=PatternFill("solid",fgColor="F8C9C9"); c0.font=Font(bold=True,color="B00000")
        if ("北斗の拳" in r["機種"] and "転生" not in r["機種"] and "無双" not in r["機種"]) or "モンキーターン" in r["機種"] or "戦国乙女" in r["機種"]:
            ws.cell(rr,1).fill=PatternFill("solid",fgColor="FFF2CC")
    for i,w in enumerate([30,11,14,9,10,9,9,9,9,13,9,9,9,9,9,11],1):
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="A3"
    wb.save("ai収集/分析_新台診断_v0.3.xlsx")
    print("\n出力: ai収集/分析_新台診断_v0.3.xlsx （データ{}機種・稼働貢献週は公式値）".format(len(recs)))
except ImportError:
    print("openpyxl無し")

print("\n=== 戦国乙女(公式稼働貢献週) ===")
for r in recs:
    if "戦国乙女" in r["機種"]:
        print(f' {r["機種"][:24]} 公式{r["稼働貢献週(公式)"]}週/設置{r["設置週数"]}週 2週持続{r["2週持続率"]} 判定={r["判定"]}')
