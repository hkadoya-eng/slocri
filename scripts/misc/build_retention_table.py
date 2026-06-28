# -*- coding: utf-8 -*-
"""
リテンション指数 比較表 v0.1（寿命平均ベース）
- 入力: ai収集/機種一覧_MY_コイン単価.csv
- 指標: リテンション指数 = SIS平均アウト ÷ コイン単価（同コストでどれだけ回させ続けたか）
- 検証: 貢献中 vs 終了 で平均リテンションに差が出るか（予測力チェック）
- 出力: ai収集/分析_スマスロ_リテンション比較_v0.1.xlsx（openpyxl無ければCSV）
注意: 真の「初週稼働値×4週持続率」は週次データが要る。これは寿命平均の暫定版。
"""
import csv, os, statistics

SRC = "ai収集/機種一覧_MY_コイン単価.csv"

def num(x):
    if x is None: return None
    s = str(x).strip().replace("+", "").replace(",", "").replace("%", "")
    if s == "": return None
    try: return float(s)
    except: return None

rows = []
with open(SRC, encoding="utf-8") as f:
    for r in csv.DictReader(f):
        rows.append(r)

recs = []
for r in rows:
    out  = num(r.get("SIS平均アウト"))
    tan  = num(r.get("コイン単価(公表)"))
    mysis= num(r.get("SIS平均MY"))
    wari = num(r.get("SIS割数%"))
    junz = num(r.get("純増(最高)"))
    atts = num(r.get("ATTS"))
    ret  = round(out / tan) if (out and tan) else None
    recs.append({
        "機種名": r.get("機種名",""),
        "メーカー": r.get("メーカー",""),
        "種類": r.get("種類",""),
        "導入": r.get("導入",""),
        "稼働週": r.get("稼働週",""),
        "状況": r.get("状況",""),
        "コイン単価": tan,
        "MY設1": num(r.get("MY(設定1)")),
        "純増": junz,
        "ATTS": atts,
        "SIS平均アウト": out,
        "SIS平均MY": mysis,
        "SIS割数%": wari,
        "リテンション指数": ret,
    })

# 計算可能なものだけ
calc = [x for x in recs if x["リテンション指数"] is not None]
calc.sort(key=lambda x: x["リテンション指数"], reverse=True)

# --- 検証: 貢献中 vs 終了 ---
def mean_ret(status_key):
    vals = [x["リテンション指数"] for x in calc if status_key in (x["状況"] or "")]
    return (round(statistics.mean(vals)), len(vals)) if vals else (None, 0)

m_live, n_live = mean_ret("貢献中")
m_dead, n_dead = mean_ret("終了")

print("=== 検証: リテンション指数の予測力 ===")
print(f"貢献中(生存) 平均リテンション = {m_live}  (n={n_live})")
print(f"終了(撤去)   平均リテンション = {m_dead}  (n={n_dead})")
if m_live and m_dead:
    print(f"→ 生存が撤去より {round((m_live/m_dead-1)*100)}% 高い" if m_live>m_dead else "→ 差が逆/小さい(指標弱い)")

# --- 北斗/モンキー ハイライト ---
print("\n=== 北斗 vs モンキー ===")
for x in calc:
    if ("北斗" in x["機種名"] or "モンキー" in x["機種名"]):
        print(f'{x["機種名"]:<28} 単価{x["コイン単価"]} MY設1{x["MY設1"]} 純増{x["純増"]} '
              f'アウト{x["SIS平均アウト"]} リテンション{x["リテンション指数"]} [{x["状況"]}]')

# --- 出力 ---
cols = ["機種名","メーカー","種類","導入","稼働週","状況","コイン単価","MY設1",
        "純増","ATTS","SIS平均アウト","SIS平均MY","SIS割数%","リテンション指数"]
out_xlsx = "ai収集/分析_スマスロ_リテンション比較_v0.1.xlsx"
out_csv  = "ai収集/分析_スマスロ_リテンション比較_v0.1.csv"

wrote = None
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook(); ws = wb.active; ws.title = "リテンション比較v0.1"
    # 注記行
    note = ("v0.1 暫定: リテンション指数=SIS平均アウト/コイン単価(寿命平均ベース)。"
            "真の初週稼働値×4週持続率は週次データが必要。台数(設置規模)の歪みは未補正。")
    ws.append([note]); ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(cols))
    ws.cell(1,1).font = Font(italic=True, color="996600")
    ws.append(cols)
    hr = ws.max_row
    for c in range(1,len(cols)+1):
        cell = ws.cell(hr,c); cell.font=Font(bold=True,color="FFFFFF")
        cell.fill=PatternFill("solid",fgColor="444444"); cell.alignment=Alignment(horizontal="center")
    for x in calc:
        ws.append([x[c] for c in cols])
        rr = ws.max_row
        nm = x["機種名"]
        if "北斗" in nm or "モンキー" in nm:
            for c in range(1,len(cols)+1):
                ws.cell(rr,c).fill = PatternFill("solid",fgColor="FFF2CC")
    # 列幅
    widths=[30,12,10,10,8,8,10,8,7,7,13,12,10,14]
    from openpyxl.utils import get_column_letter
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="A3"
    wb.save(out_xlsx); wrote=out_xlsx
except ImportError:
    with open(out_csv,"w",encoding="utf-8-sig",newline="") as f:
        w=csv.writer(f); w.writerow(cols)
        for x in calc: w.writerow([x[c] for c in cols])
    wrote=out_csv

print(f"\n出力: {wrote}  (計算可能 {len(calc)} 機種 / 全 {len(recs)} 行)")
