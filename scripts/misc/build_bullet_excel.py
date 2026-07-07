# -*- coding: utf-8 -*-
"""バレットサークル 気持ちよさ分解＋派生提案(照準チャージ型) をExcel化。"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
wb=Workbook()
H=PatternFill("solid",fgColor="444444"); HF=Font(bold=True,color="FFFFFF")
def hdr(ws,row,cols):
    for i,c in enumerate(cols,1):
        x=ws.cell(row,i,c); x.font=HF; x.fill=H; x.alignment=Alignment(horizontal="center")

# ---- 気持ちよさ分解 ----
ws=wb.active; ws.title="気持ちよさ分解"
ws.column_dimensions["A"].width=22; ws.column_dimensions["B"].width=70
ws.cell(1,1,"バレットサークル(SAO2) 気持ちよさ／癖になる表現の分解").font=Font(bold=True,color="D85A30",size=14)
ws.cell(2,1,"※擬似遊技でなくリール上で図柄を止める『純粋なヒキ＋狙い』が主役。実戦の声・解析(一撃/なな徹/スロ板RUSH)より").font=Font(italic=True,color="888888")
hdr(ws,3,["気持ちよさの柱","中身"])
r=4
for a,b in [
 ("①主体感(擬似遊技でない)","枠に図柄が止まるかは目押し次第。押した瞬間に決まる=出玉が自分の手の中。液晶の当落待ちの受け身と真逆＝快感の源泉"),
 ("②身体性","銃の役物ヘカートを構えスコープを覗く物理動作。撃つ・狙うの体を使う操作が没入と手応えを深める"),
 ("③即時フィードバック","枠にカチッと止まった瞬間の視覚・音・役物が即返る。狙って当てた達成感がその場で戻るループ"),
 ("④上達の快感","左リール緑7の2連狙いで1確目を先読み。打つほど察知が上手くなり熟練者ほど優越感"),
 ("⑤惜しさ(near-miss)","枠にあと1コマで止まらない悔しさが『もう一度』を生む。中毒性を静かに支える"),
 ("⑥期待の波","毎G枠(1〜6個・13パターン)が変わり期待が更新され続ける＝中だるみしない"),
]:
    ws.cell(r,1,a); c=ws.cell(r,2,b); c.alignment=Alignment(wrap_text=True); r+=1

r+=1
ws.cell(r,1,"■ 癖になる表現の設計原理(game-feel)").font=Font(bold=True,color="D85A30"); r+=1
hdr(ws,r,["原理","バレットサークルでの現れ"]); r+=1
for a,b in [
 ("主体感(agency)","自分の目押しで結果が変わる"),
 ("即時報酬","止めた瞬間に視覚/音/役物が返す"),
 ("near-miss","あと1コマで外す→もう一回引きたい"),
 ("mastery(上達)","狙い・1確察知で腕が上がる"),
 ("身体性","銃を構える/スコープを覗く物理役物"),
 ("期待の波","毎G枠が変わり期待が途切れない"),
]:
    ws.cell(r,1,a); ws.cell(r,2,b); r+=1

r+=1
ws.cell(r,1,"■ 派生への示唆(気持ちよさを増幅)").font=Font(bold=True,color="D85A30"); r+=1
for s in ["止めた瞬間の報酬を段階的に派手化＝即時報酬を最大化","わざと near-miss を見せ次Gの期待を煽る","狙う枠を打ち手が選べる＝主体感をさらに前面に","1確察知・狙い精度で期待値が変わる＝上達を出玉に直結"]:
    c=ws.cell(r,1,"・"+s); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2); r+=1

# ---- 派生提案 ----
ws2=wb.create_sheet("派生提案 照準チャージ型")
ws2.column_dimensions["A"].width=16
for col in "BCDEF": ws2.column_dimensions[col].width=15
ws2.cell(1,1,"派生ゲーム性提案『照準チャージ型』(オリジナル試算・実機非紐付け)").font=Font(bold=True,color="D85A30",size=13)
rows=[
 ("設計思想","バレットサークルの『自分で止める快感』を、生死を決める持続率×稼働値に接続。気持ちいいから打ち続ける台を狙う"),
 ("通常時","リールにターゲット枠→止めてチャージ(自力)。狙う枠を選べる=主体感拡張。near-missで惜しさ→継続動機。MAXでCZ"),
 ("AT(照準ラッシュ)","枠に止めるたび継続ゲージ/差枚を自力上乗せ(即時報酬)。ヒット連鎖で上位ATへ自力突入"),
 ("やめ時消去","AT後もターゲット高確が継続、止められる限りループ"),
 ("上達→出玉","1確察知・狙い精度で期待値が変わる=技術介入を報酬化"),
]
r=3
for a,b in rows:
    ws2.cell(r,1,a).font=Font(bold=True); c=ws2.cell(r,2,b); c.alignment=Alignment(wrap_text=True); ws2.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6); r+=1
r+=1
ws2.cell(r,1,"■ 数値設計＋自己検算").font=Font(bold=True,color="1f9d4d"); r+=1
hdr(ws2,r,["区分","純増","1セットG","継続率","平均連(1/(1-継続))","平均出玉(純増×G×連)"]); r+=1
for a,b,c,d,e,f in [("メインAT",3.0,40,"70%","約3.33連","約400枚"),("上位AT",6.0,40,"85%","約6.67連","約1600枚")]:
    for i,v in enumerate([a,b,c,d,e,f],1): ws2.cell(r,i,v)
    r+=1
for s in ["初当り約1/300／コイン単価約3.5円／機械割 約97.5〜114.5%（現実レンジ内）","自己検算：平均連1/(1-0.70)=3.33・1/(1-0.85)=6.67 ✓／平均出玉 3.0×40×3.33≒400・6.0×40×6.67≒1600 ✓","特化ゾーン『連射チャンス』＝枠ヒット連鎖で止めるほど加速する即時報酬の塊","正直な穴：流行りそうは仮説。供給過多・試験の運で死ぬ。最終判定は打ち手の審美眼＝打って痺れるか"]:
    c=ws2.cell(r,1,s); ws2.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); c.alignment=Alignment(wrap_text=True); r+=1

wb.save("ai収集/分析_バレットサークル気持ちよさ＋派生照準チャージ型_v1.xlsx")
print("出力: ai収集/分析_バレットサークル気持ちよさ＋派生照準チャージ型_v1.xlsx")
