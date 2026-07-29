import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "ゲーム性一覧"

C_HEADER  = "1F3864"
C_NORMAL  = "D9EAD3"
C_CZ      = "FCE5CD"
C_AT      = "CFE2F3"
C_BONUS   = "EAD1DC"
C_COMPLEX = "FFF2CC"
C_SUB     = "F3F3F3"
WHITE = "FFFFFF"
BLACK = "000000"

def cell_style(ws, row, col, value, bg, font_color=BLACK, bold=False, size=10, wrap=False, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(name="Yu Gothic UI", color=font_color, bold=bold, size=size)
    ha = "center" if align == "center" else "left"
    c.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=wrap)
    c.border = Border(
        left=Side(style="thin", color="AAAAAA"),
        right=Side(style="thin", color="AAAAAA"),
        top=Side(style="thin", color="AAAAAA"),
        bottom=Side(style="thin", color="AAAAAA"),
    )
    return c

# タイトル
ws.merge_cells("A1:G1")
t = ws["A1"]
t.value = "パチスロ ゲーム性分類一覧"
t.fill = PatternFill("solid", fgColor=C_HEADER)
t.font = Font(name="Yu Gothic UI", color=WHITE, bold=True, size=14)
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 32

# ヘッダー
headers = ["大カテゴリ", "型名", "純増", "ゲーム性の核心", "代表的な設計パターン", "やめどき/特徴", "代表機種例"]
for i, h in enumerate(headers, 1):
    cell_style(ws, 2, i, h, C_HEADER, WHITE, bold=True, size=10, align="center")
ws.row_dimensions[2].height = 22

# データ: (大カテゴリ, 型名, 純増, 核心, パターン, やめどき, 機種例, 色)
rows = [
    ("通常ゲーム性", "周期抽選型", "なし",
     "一定周期ごとにCZ/AT抽選を実施。周期終了時に必ずチャンス",
     "周期50〜100G。周期内の小役でポイント蓄積→終了時抽選",
     "周期抜け・CZ非突入なら即やめ候補",
     "バジリスク絆2, スマスロ天元突破G", C_NORMAL),
    ("通常ゲーム性", "ゾーン型", "なし",
     "特定G数帯でAT/CZ当選確率が大幅上昇するゾーンが存在",
     "100G・200G・300G前後がゾーン帯。ゾーン外はスルー確認推奨",
     "ゾーン抜け後・前兆なければやめ",
     "番長3, まどマギ2", C_NORMAL),
    ("通常ゲーム性", "天井型", "なし",
     "規定G数消化で必ずCZ/AT/ボーナスが当たる救済機能",
     "天井G数は機種ごとに設定。設定変更でリセットされる場合も",
     "天井到達後・AT終了後は即やめ",
     "多数（汎用設計）", C_NORMAL),
    ("通常ゲーム性", "前兆演出型", "なし",
     "内部当選後、演出（前兆）を経てからCZ/ATに突入する設計",
     "前兆中は特定小役・演出で盛り上がる。即告知型と対比",
     "前兆終了・非当選確認後やめ",
     "Re:ゼロ, 鬼滅の刃", C_NORMAL),
    ("通常ゲーム性", "Aタイプ型", "なし",
     "ボーナス（BB/RB）成立のみで完結するシンプル設計",
     "出玉=ボーナス枚数のみ。設定差が純粋に出玉に直結",
     "ボーナス後即やめOK",
     "ジャグラー系, ハナハナ系", C_NORMAL),
    ("通常ゲーム性", "ポイント蓄積型", "なし",
     "通常時に小役でポイントを貯め、一定量でCZ/AT抽選",
     "メーター等でポイントを視覚化。引き継ぎ確認が重要",
     "ポイント少量・前兆なしでやめ",
     "北斗の拳 天昇, 押忍!番長4", C_NORMAL),

    ("CZ（チャレンジゾーン）", "固定ゲーム数型", "なし",
     "規定G数内にAT当選条件を満たせば突入。時間が決まっている",
     "30G・50G・100Gなど固定。消化中の小役引きが鍵",
     "CZ失敗→天井狙いか即やめ判断",
     "多数", C_CZ),
    ("CZ（チャレンジゾーン）", "継続率型（セット型）", "なし",
     "毎セット終了時にCZ継続orAT突入を抽選。継続確率が設定に依存",
     "高設定ほど継続率UP。セット継続演出が豊富",
     "CZ終了・AT非突入でやめ",
     "番長シリーズ, 沖ドキ系", C_CZ),
    ("CZ（チャレンジゾーン）", "バトル型（敵撃破型）", "なし",
     "敵（ボス）とバトルして勝利でAT突入。負けると終了",
     "敵HPを削る演出。勝率は設定や突入経緯で変化",
     "CZバトル敗北→状況次第でやめ",
     "北斗の拳系, ゴッドシリーズ", C_CZ),
    ("CZ（チャレンジゾーン）", "成否二択型", "なし",
     "演出の最後に成功か失敗かで結果が決まるシンプル構造",
     "最後の一発演出で判明。前兆演出との組み合わせが多い",
     "失敗→天井・ゾーンで判断",
     "汎用（多機種）", C_CZ),
    ("CZ（チャレンジゾーン）", "昇格チャレンジ型", "なし",
     "複数段階のランク/レベルがあり、上位に昇格するほどAT当選率UP",
     "ブロンズ→シルバー→ゴールドと昇格。段階演出が豊富",
     "最上位失敗→やめ判断",
     "まどマギ叛逆, Re:ゼロ", C_CZ),
    ("CZ（チャレンジゾーン）", "ミッション達成型", "なし",
     "特定ミッション（特定役成立等）を達成するとAT突入",
     "ミッション内容が事前開示されることも。達成難易度で期待度変化",
     "ミッション失敗→即やめ候補",
     "一部機種", C_CZ),
    ("CZ（チャレンジゾーン）", "ルーレット/選択型", "なし",
     "複数の選択肢からランダム or 演出でAT/外れを決定",
     "見た目は選択式だが内部は事前決定。結果演出の緊張感が売り",
     "外れ→やめ",
     "一部機種", C_CZ),

    ("AT（アシストタイム）", "固定ゲーム数型", "あり",
     "規定G数を消化すればAT終了。枚数でなくG数管理",
     "50G・100G・200G固定。消化中の上乗せ抽選がある機種も",
     "AT終了後→天井狙いか即やめ",
     "多数", C_AT),
    ("AT（アシストタイム）", "セット継続型（継続率型）", "あり",
     "毎セット終了時に継続抽選。継続率が機種の荒さを決める主要因",
     "継続率50〜80%が多い。ラスト演出→継続ガチ抽選の流れが定番",
     "AT終了・前兆なし→やめ",
     "バジリスク絆, 沖ドキDUO", C_AT),
    ("AT（アシストタイム）", "ゲーム数上乗せ型", "あり",
     "AT中に特定役でゲーム数を上乗せしてATを延長",
     "上乗せ量が数G〜数百Gで幅広い。上乗せ特化ゾーンが派生する機種も",
     "上乗せなく残G少→準備",
     "Re:ゼロ, 天元突破G", C_AT),
    ("AT（アシストタイム）", "差枚数管理型", "あり",
     "ATは規定枚数（差枚数）到達で終了。G数ではなく出玉で管理",
     "1500枚・3000枚上限が多い。スマスロに多い設計",
     "上限枚数到達→終了",
     "スマスロ系全般", C_AT),
    ("AT（アシストタイム）", "ループ型", "あり",
     "AT終了後に一定確率で再突入（ループ）する設計",
     "ループ率50〜90%程度。連荘が続くほど出玉爆増の可能性",
     "ループ外れ→即やめ",
     "ハーデス, ミリオンゴッド", C_AT),
    ("AT（アシストタイム）", "レベルアップ型", "あり",
     "AT中にレベルが上昇し、上位レベルほど純増や恩恵が強化",
     "Lv1→Lv3で純増3枚→5枚など。Lv達成演出が盛り上がりポイント",
     "最高Lv終了→状況次第",
     "花火絶景, 一部新台", C_AT),
    ("AT（アシストタイム）", "AT内CZ型（階層型）", "あり",
     "AT中にさらにCZが発生し、上位ATや特化ゾーンへ昇格する多層構造",
     "通常AT→CZ（昇格抽選）→上位ATの流れ。恩恵の段階が多い",
     "上位AT終了→判断",
     "北斗の拳 天昇", C_AT),
    ("AT（アシストタイム）", "特化ゾーン型", "あり",
     "AT中に発生する超高純増の特化ゾーン（上乗せ特化）がメイン出玉源",
     "特化ゾーン=爆発の本番。ゾーン中の引きで出玉が大きく変わる",
     "特化ゾーン終了→残G確認",
     "まどマギ, 鬼滅の刃", C_AT),

    ("ボーナス型", "BB（ビッグボーナス）型", "あり（一時的）",
     "規定枚数（240〜315枚程度）を獲得するボーナス。旧来設計",
     "旧来の4〜5号機設計。技術介入（リプレイはずし）要素があった機種も",
     "BB終了→即やめ可（純Aなら）",
     "初代北斗の拳, 吉宗（4号機）", C_BONUS),
    ("ボーナス型", "ST（ストック）ボーナス型", "あり",
     "ボーナスをストックし、放出タイミングを内部抽選で管理",
     "ストック機はやめどきが難しい。ハマりとドカ連が交互に来る設計",
     "連荘終了→前兆確認後やめ",
     "ミリオンゴッド（初代）", C_BONUS),

    ("複合型", "CZ+AT複合型", "CZなし/ATあり",
     "通常→CZ（純増なし）→AT（純増あり）の2段階構造が明確な設計",
     "CZは橋渡し役。AT突入してから本番。CZ失敗が最大のストレスポイント",
     "CZ失敗→天井G数で判断",
     "スマスロ系（多数）", C_COMPLEX),
    ("複合型", "周期+CZ+AT型", "CZなし/ATあり",
     "周期で抽選→CZ突入→AT突入の3段階。各段階でドラマが生まれる",
     "天井も絡むと4段階に。設計が複雑なほど演出の幅が広い",
     "周期通過・CZ失敗→天井確認",
     "バジリスク絆2系", C_COMPLEX),
    ("複合型", "AT中ループ+上乗せ複合型", "あり",
     "ATが継続率×上乗せのダブル管理。どちらかが切れたら終了",
     "継続率と上乗せGの両方を管理。終了演出が複雑になりがち",
     "どちらか尽きたら終了",
     "一部スマスロ新機種", C_COMPLEX),
]

row_num = 3
for r in rows:
    cat, name, junie, core, pattern, stop, example, color = r
    cell_style(ws, row_num, 1, cat, color, bold=True, wrap=True, align="center")
    cell_style(ws, row_num, 2, name, color, bold=True)
    cell_style(ws, row_num, 3, junie, color, align="center")
    cell_style(ws, row_num, 4, core, color, wrap=True)
    cell_style(ws, row_num, 5, pattern, color, wrap=True)
    cell_style(ws, row_num, 6, stop, color, wrap=True)
    cell_style(ws, row_num, 7, example, color, wrap=True)
    ws.row_dimensions[row_num].height = 52
    row_num += 1

ws.column_dimensions["A"].width = 16
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 10
ws.column_dimensions["D"].width = 30
ws.column_dimensions["E"].width = 36
ws.column_dimensions["F"].width = 22
ws.column_dimensions["G"].width = 24
ws.freeze_panes = "A3"

# ===== Sheet 2: 凡例 =====
ws2 = wb.create_sheet("凡例・見方")
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 55

ws2.merge_cells("A1:B1")
h = ws2["A1"]
h.value = "カテゴリ カラー凡例"
h.fill = PatternFill("solid", fgColor=C_HEADER)
h.font = Font(name="Yu Gothic UI", color=WHITE, bold=True, size=12)
h.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28

legends = [
    (C_NORMAL,  "通常ゲーム性",          "ボーナス/CZ/AT突入前の通常時フロー。天井・ゾーン・周期などが該当"),
    (C_CZ,      "CZ（チャレンジゾーン）", "純増なし。AT突入を目指す橋渡しゾーン。失敗すると通常に戻る"),
    (C_AT,      "AT（アシストタイム）",   "純増あり。これが本番の出玉獲得フェーズ。セット・G数・枚数で管理"),
    (C_BONUS,   "ボーナス型",            "旧来BB/RBを主体とした設計。4〜5号機中心"),
    (C_COMPLEX, "複合型",               "複数カテゴリを組み合わせた多層設計"),
]
for i, (color, name, desc) in enumerate(legends, 2):
    c1 = ws2.cell(row=i, column=1, value=name)
    c1.fill = PatternFill("solid", fgColor=color)
    c1.font = Font(name="Yu Gothic UI", bold=True, size=10)
    c1.alignment = Alignment(vertical="center")
    ws2.row_dimensions[i].height = 22
    c2 = ws2.cell(row=i, column=2, value=desc)
    c2.fill = PatternFill("solid", fgColor=color)
    c2.font = Font(name="Yu Gothic UI", size=10)
    c2.alignment = Alignment(vertical="center", wrap_text=True)

r = len(legends) + 3
ws2.merge_cells(f"A{r}:B{r}")
t2 = ws2[f"A{r}"]
t2.value = "列の説明"
t2.fill = PatternFill("solid", fgColor=C_HEADER)
t2.font = Font(name="Yu Gothic UI", color=WHITE, bold=True, size=11)
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[r].height = 24

col_desc = [
    ("大カテゴリ", "通常/CZ/AT/ボーナス/複合 の5分類"),
    ("型名", "◯◯型という命名で設計パターンを識別"),
    ("純増", "そのゾーン内で出玉が増えるか（あり=AT・なし=CZ/通常）"),
    ("ゲーム性の核心", "その型が持つ最大の特徴・仕組みの説明"),
    ("代表的な設計パターン", "実際の数値・演出構成の具体例"),
    ("やめどき/特徴", "立ち回り上の注意点・やめどきの目安"),
    ("代表機種例", "この型に当てはまる代表的なパチスロ機種"),
]
for j, (col, desc) in enumerate(col_desc, r + 1):
    c1 = ws2.cell(row=j, column=1, value=col)
    c1.fill = PatternFill("solid", fgColor=C_SUB)
    c1.font = Font(name="Yu Gothic UI", bold=True, size=10)
    c1.alignment = Alignment(vertical="center")
    ws2.row_dimensions[j].height = 20
    c2 = ws2.cell(row=j, column=2, value=desc)
    c2.font = Font(name="Yu Gothic UI", size=10)
    c2.alignment = Alignment(vertical="center")

path = r"C:\Users\h.kadoya\Desktop\slocri\proposals\ゲーム性分類一覧.xlsx"
wb.save(path)
print(f"完了: {path}")
print(f"合計: {len(rows)} 型")
