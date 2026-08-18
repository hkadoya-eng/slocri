"""
週毎SISデータ一覧_2026.xlsm から稼働貢献週を Supabase の sis_machine_stats にインポートする。

使い方:
  python import_sis_weekly.py
"""

import sys
import io
import os
import re
import json
import requests

ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

WEEKLY_PATH = "Z:/01_SISデータ/PS/週毎SISデータ一覧_2026.xlsm"
SUPABASE_URL = "https://vpzbtuucopucablwyqeq.supabase.co"

# ファイル名末尾から年を抽出 (例: 週毎SISデータ一覧_2026.xlsm → 2026)
_year_match = re.search(r"_(\d{4})\.xlsm", WEEKLY_PATH)
FILE_YEAR = int(_year_match.group(1)) if _year_match else 2026


def split_month_day(part, start_month=None):
    """'5.3' → (5, 3)。区切りが抜けた '713' '720' のような表記も許容する。
       (シート名 '7.7~713' '7.14~720' が実在し、弾くとその週が丸ごと欠落する)
       曖昧さは「終了月は開始月 or 翌月」という週シートの性質で解消する。"""
    part = part.strip().lstrip(".")
    ps = part.split(".")
    if len(ps) == 2:
        try:
            return int(ps[0]), int(ps[1])
        except ValueError:
            return None
    if len(ps) != 1 or not part.isdigit() or not (3 <= len(part) <= 4):
        return None
    cands = []
    if start_month is not None:
        cands = [start_month, 1 if start_month == 12 else start_month + 1]
    else:
        cands = list(range(1, 13))
    for m in cands:
        pre = str(m)
        if part.startswith(pre):
            d_str = part[len(pre):]
            if d_str.isdigit() and 1 <= int(d_str) <= 31:
                return m, int(d_str)
    return None


def parse_week_range(s):
    """'4.27~5.3' → '2026-04-27' / '12.26~1.1' → '2025-12-26' (年跨ぎ補正あり)"""
    if not isinstance(s, str) or "~" not in s:
        return None
    start_part, end_part = s.split("~", 1)
    sp = split_month_day(start_part)
    if not sp:
        return None
    ep = split_month_day(end_part, sp[0])
    if not ep:
        return None
    sm, sd = sp
    em = ep[0]
    if not (1 <= sm <= 12 and 1 <= sd <= 31 and 1 <= em <= 12):
        return None
    # 年跨ぎ判定: 開始月 > 終了月 (例: 12→1) なら開始は前年扱い
    year = FILE_YEAR - 1 if sm > em else FILE_YEAR
    return f"{year:04d}-{sm:02d}-{sd:02d}"


def load_env_local():
    env_path = os.path.join(ROOT_DIR, ".env.local")
    if not os.path.exists(env_path):
        return
    with open(env_path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                key, _, val = line.partition("=")
                os.environ.setdefault(key.strip(), val.strip())


def extract_records(path):
    try:
        import openpyxl
    except ImportError:
        print("openpyxl が未インストールです: pip install openpyxl")
        sys.exit(1)

    print(f"Excelを読み込み中: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=True)

    ws = wb["機種一覧表"]
    rows = list(ws.iter_rows(values_only=True))

    # ヘッダー行1のcol3が最新確定週（例: "4.27~5.3"）
    header = rows[1]
    week_str = header[3] if len(header) > 3 else None
    last_week_start = parse_week_range(week_str)

    records = []
    for row in rows[2:]:  # row 0〜1はヘッダー
        machine = row[1]
        contrib_raw = row[2]
        if not machine or not isinstance(machine, str):
            continue
        if not machine.startswith("L"):
            continue
        # "X週稼動貢献中" からX を抽出
        if contrib_raw and isinstance(contrib_raw, str):
            m = re.search(r"(\d+)週", contrib_raw)
            weeks = int(m.group(1)) if m else 0
        else:
            weeks = 0

        records.append({
            "machine": machine.strip(),
            "contrib_weeks": weeks,
        })

    wb.close()
    return records, last_week_start


def upsert_records(records, api_key):
    headers = {
        "apikey": api_key,
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates",
    }

    total = len(records)
    res = requests.post(
        f"{SUPABASE_URL}/rest/v1/sis_machine_stats",
        headers=headers,
        data=json.dumps(records),
    )
    if res.status_code in (200, 201):
        print(f"  アップロード完了: {total} 件")
    else:
        print(f"  エラー ({res.status_code}): {res.text[:200]}")


def parse_week_start(sheet_name):
    """シート名から週開始日(月曜)を返す。年跨ぎは前年扱い。例: '4.27~5.3' → '2026-04-27' / '12.26~1.1' → '2025-12-26'"""
    return parse_week_range(sheet_name)


def build_sheet_year_map(sheet_names):
    """シート順に年境界(月の減少)を数え、各シート名→正しいyyyy-mm-dd を返す。
       最後の週シートが FILE_YEAR。"""
    parsed = []
    for s in sheet_names:
        if "~" not in s:
            continue
        sp_raw, ep_raw = s.split("~", 1)
        sp = split_month_day(sp_raw)
        if not sp:
            continue
        ep = split_month_day(ep_raw, sp[0])
        if not ep:
            continue
        sm, sd = sp
        em, ed = ep
        if not (1 <= sm <= 12 and 1 <= sd <= 31 and 1 <= em <= 12 and 1 <= ed <= 31):
            continue
        # 週シートは「開始月と同じ月で開始日より後」か「翌月」で終わるはず。
        # それ以外(例: '8.27~8.2' = 7.27~8.2 のシート名タイポ)は壊れた見出しなので採用しない。
        # 採用すると後続の forward sweep が偽の年境界を検出し、全シートが1年ずれる。
        if em == sm:
            # 見出しは '6.2~6.8'(日曜終わり) と '6.2~6.9'(翌月曜終わり) の両方の書き方が混在するので +7 まで許容
            valid = sd < ed <= sd + 7
        elif em == sm + 1 or (sm == 12 and em == 1):
            valid = True
        else:
            valid = False
        if not valid:
            print(f"  [skip] 週見出しが不正なシートを除外: {s}")
            continue
        parsed.append({"sheet": s, "sm": sm, "sd": sd, "em": em})
    if not parsed:
        return {}
    # Forward sweep: 開始月が前のシートより大きく戻ったら年が変わった。
    # 12→1 の本物の年跨ぎは月が11戻るので、しきい値6で誤検出(見出しの前後入れ替わり等)を弾く。
    rel = 0
    prev_sm = parsed[0]["sm"]
    for p in parsed:
        if p["sm"] < prev_sm - 6:
            rel += 1
        p["rel"] = rel
        prev_sm = p["sm"]
    max_rel = max(p["rel"] for p in parsed)
    base_year = FILE_YEAR - max_rel
    mapping = {}
    for p in parsed:
        year = base_year + p["rel"]
        mapping[p["sheet"]] = f"{year:04d}-{p['sm']:02d}-{p['sd']:02d}"
    return mapping


def extract_weekly_records(path):
    try:
        import openpyxl
    except ImportError:
        print("openpyxl が未インストールです")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=True)
    sheet_year_map = build_sheet_year_map(wb.sheetnames)
    all_records = []
    for sheet_name in wb.sheetnames:
        week_start = sheet_year_map.get(sheet_name)
        if not week_start:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        for row in rows[1:]:
            machine = row[1]
            if not machine or not isinstance(machine, str) or not machine.startswith("L"):
                continue
            out = row[6]
            if not isinstance(out, (int, float)):
                continue
            profit = row[8]
            payout = row[11]
            coin_price = row[9]
            avg_count = row[5]
            all_records.append({
                "machine": machine.strip(),
                "week_start": week_start,
                "out_coins": float(out),
                "gross_profit": float(profit) if isinstance(profit, (int, float)) else None,
                "payout_rate": float(payout) if isinstance(payout, (int, float)) else None,
                "coin_price": float(coin_price) if isinstance(coin_price, (int, float)) else None,
                "avg_machine_count": float(avg_count) if isinstance(avg_count, (int, float)) else None,
            })
    wb.close()
    return all_records


def upsert_weekly_records(records, api_key):
    if not records:
        return
    headers = {
        "apikey": api_key,
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates",
    }
    chunk_size = 500
    total = 0
    for i in range(0, len(records), chunk_size):
        chunk = records[i:i + chunk_size]
        res = requests.post(
            f"{SUPABASE_URL}/rest/v1/sis_weekly_data",
            headers=headers,
            data=json.dumps(chunk),
        )
        if res.status_code in (200, 201):
            total += len(chunk)
        else:
            print(f"  エラー ({res.status_code}): {res.text[:200]}")
    print(f"  週次データアップロード完了: {total} 件")


def upsert_config(last_week_start, api_key):
    if not last_week_start:
        return
    headers = {
        "apikey": api_key,
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates",
    }
    config = [{"machine": "__config__", "contrib_weeks": 0, "last_week_start": last_week_start}]
    res = requests.post(
        f"{SUPABASE_URL}/rest/v1/sis_machine_stats",
        headers=headers,
        data=json.dumps(config),
    )
    if res.status_code in (200, 201):
        print(f"  最終確定週を保存: {last_week_start}")
    else:
        print(f"  configエラー ({res.status_code}): {res.text[:200]}")


def run():
    load_env_local()

    api_key = os.environ.get("SUPABASE_SERVICE_KEY") or os.environ.get("VITE_SUPABASE_ANON_KEY")
    if not api_key:
        print("エラー: VITE_SUPABASE_ANON_KEY が未設定です")
        sys.exit(1)

    win_path = WEEKLY_PATH.replace("/", "\\")
    if not os.path.exists(win_path):
        print(f"エラー: ファイルが見つかりません: {WEEKLY_PATH}")
        sys.exit(1)

    records, last_week_start = extract_records(WEEKLY_PATH)
    print(f"\n抽出: {len(records)} 機種　最終確定週: {last_week_start}\n")

    if not records:
        print("レコードなし。終了します。")
        return

    print("Supabase にアップロード中...")
    upsert_records(records, api_key)
    upsert_config(last_week_start, api_key)

    print("\n週次データを抽出中...")
    weekly_records = extract_weekly_records(WEEKLY_PATH)
    print(f"  {len(weekly_records)} 件")
    upsert_weekly_records(weekly_records, api_key)
    print("\n完了。")

    print("稼働貢献週 上位10機種:")
    top = sorted(records, key=lambda r: r["contrib_weeks"], reverse=True)[:10]
    for r in top:
        print(f"  {r['contrib_weeks']:3d}週  {r['machine']}")


if __name__ == "__main__":
    run()
