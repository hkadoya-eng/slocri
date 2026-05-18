"""
SISデータ（稼働まとめ.xlsm）を Supabase の sis_data テーブルにインポートする。

使い方:
  python import_sis.py [--force]

  --force: 既存レコードも上書きupsert（デフォルトはonConflict=ignoreで新規のみ追加）

必要な環境変数（.env.local）:
  VITE_SUPABASE_ANON_KEY=eyJ...  （通常の anon キー。SERVICE_KEY 不要）
"""

import sys
import io
import os
import json
import requests
from datetime import datetime, date

ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

SIS_PATH = "Z:/01_SISデータ/PS/PS日毎稼働まとめ_2026.xlsm"
SUPABASE_URL = "https://vpzbtuucopucablwyqeq.supabase.co"

FORCE = "--force" in sys.argv


def load_env_local():
    env_path = os.path.join(ROOT_DIR, ".env.local")
    if not os.path.exists(env_path):
        return
    with open(env_path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                key, _, val = line.partition("=")
                os.environ.setdefault(key.strip(), val.strip())


def extract_records(path):
    try:
        import openpyxl
    except ImportError:
        print("openpyxl が未インストールです: pip install openpyxl")
        sys.exit(1)

    print(f"Excelを読み込み中: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=True)

    # 週シートのみ対象（260427～ のような名前）
    target_sheets = [s for s in wb.sheetnames if s[:2].isdigit()]
    print(f"対象シート: {target_sheets}")

    records = []
    for sheet_name in target_sheets:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 10:
            continue

        dates = rows[2][6:13]  # G〜M列 = 月〜日の日付

        i = 4  # 全国アウト行をスキップして開始
        while i < len(rows) - 5:
            row = rows[i]
            machine = row[1]
            # L/Sで始まる機種名（【】を含まないもの）。S系=5号機(ジャグラー/ハナハナ等)も含める
            if machine and isinstance(machine, str) and (machine.startswith("L") or machine.startswith("S")) and "【" not in machine:
                out_vals    = rows[i][6:13]
                coin_vals   = rows[i + 1][6:13]
                rate_vals   = rows[i + 2][6:13]
                profit_vals = rows[i + 3][6:13]
                ratio_vals  = rows[i + 4][6:13]
                count_vals  = rows[i + 5][6:13]

                for j, d in enumerate(dates):
                    if isinstance(d, datetime) and isinstance(out_vals[j], (int, float)):
                        records.append({
                            "machine": machine.strip(),
                            "date": d.strftime("%Y-%m-%d"),
                            "out_coins": int(out_vals[j]),
                            "coin_price": float(coin_vals[j]) if isinstance(coin_vals[j], (int, float)) else None,
                            "payout_rate": float(rate_vals[j]) if isinstance(rate_vals[j], (int, float)) else None,
                            "gross_profit": int(profit_vals[j]) if isinstance(profit_vals[j], (int, float)) else None,
                            "operation_ratio": float(ratio_vals[j]) if isinstance(ratio_vals[j], (int, float)) else None,
                            "machine_count": int(count_vals[j]) if isinstance(count_vals[j], (int, float)) else None,
                        })
                i += 6
            else:
                i += 1

    wb.close()
    return records


def aggregate_national_payout_rate(records, service_key):
    """機種別データから日付ごとの全国出玉率（IN加重平均）を計算してsis_national_dailyに反映する。
    SISのExcelには全国出玉率の直接欄が無いため、機種別出玉率(枚/枚)をout_coinsで加重平均する。"""
    by_date = {}
    for r in records:
        if r.get("payout_rate") is None or not r.get("out_coins"):
            continue
        d = r["date"]
        by_date.setdefault(d, [0.0, 0.0])
        by_date[d][0] += r["out_coins"] * r["payout_rate"]
        by_date[d][1] += r["out_coins"]

    if not by_date:
        return 0

    payload = []
    for d, (sumWP, sumW) in by_date.items():
        if sumW > 0:
            payload.append({"date": d, "payout_rate": round(sumWP / sumW, 2)})

    if not payload:
        return 0

    headers = {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates",
    }
    BATCH = 500
    updated = 0
    for i in range(0, len(payload), BATCH):
        batch = payload[i:i + BATCH]
        res = requests.post(
            f"{SUPABASE_URL}/rest/v1/sis_national_daily?on_conflict=date",
            headers=headers,
            data=json.dumps(batch),
        )
        if res.status_code in (200, 201):
            updated += len(batch)
        else:
            print(f"  全国出玉率エラー ({res.status_code}): {res.text[:200]}")
    return updated


def upsert_records(records, service_key):
    headers = {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates" if FORCE else "resolution=ignore-duplicates",
    }

    BATCH = 200
    total = len(records)
    inserted = 0

    for i in range(0, total, BATCH):
        batch = records[i:i + BATCH]
        res = requests.post(
            f"{SUPABASE_URL}/rest/v1/sis_data?on_conflict=machine,date",
            headers=headers,
            data=json.dumps(batch),
        )
        if res.status_code in (200, 201):
            inserted += len(batch)
            print(f"  アップロード {i + len(batch)}/{total} 件")
        else:
            print(f"  エラー ({res.status_code}): {res.text[:200]}")

    return inserted


def run():
    load_env_local()

    service_key = os.environ.get("SUPABASE_SERVICE_KEY") or os.environ.get("VITE_SUPABASE_ANON_KEY")
    if not service_key:
        print("エラー: VITE_SUPABASE_ANON_KEY が未設定です")
        sys.exit(1)

    if not os.path.exists(SIS_PATH.replace("/", "\\")):
        # Windows パス変換
        win_path = SIS_PATH.replace("/", "\\")
        if not os.path.exists(win_path):
            print(f"エラー: ファイルが見つかりません: {SIS_PATH}")
            print("  Google Drive for Desktop が起動中か確認してください")
            sys.exit(1)

    records = extract_records(SIS_PATH)
    print(f"\n抽出: {len(records)} 件\n")

    if not records:
        print("レコードなし。終了します。")
        return

    print(f"Supabase にアップロード中... (--force={FORCE})")
    upsert_records(records, service_key)

    # 全国出玉率を機種別データから集計してsis_national_dailyへ
    print("\n全国出玉率を集計中...")
    n = aggregate_national_payout_rate(records, service_key)
    print(f"  {n} 日分の payout_rate を更新")

    print(f"\n完了。")

    # 日付範囲を表示
    dates = sorted(set(r["date"] for r in records))
    print(f"期間: {dates[0]} 〜 {dates[-1]}")
    machines = sorted(set(r["machine"] for r in records))
    print(f"機種数: {len(machines)}")


if __name__ == "__main__":
    run()
