"""
日毎稼働全体.xlsx から全国日次集計データを Supabase の sis_national_daily テーブルにインポートする。

使い方:
  python import_national_daily.py
"""

import sys
import io
import os
import re
import json
import requests
from datetime import datetime, timedelta

ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

EXCEL_PATH = "Z:/01_SISデータ/PS/日毎稼働全体.xlsx"
WEEKLY_PATH = "Z:/01_SISデータ/PS/PS日毎稼働まとめ_2026.xlsm"
SUPABASE_URL = "https://vpzbtuucopucablwyqeq.supabase.co"
SHEET_NAME = "他機種含む"

# Excelシリアル日付 → date文字列
EXCEL_EPOCH = datetime(1899, 12, 30)


def serial_to_date(serial):
    return (EXCEL_EPOCH + timedelta(days=int(serial))).strftime("%Y-%m-%d")


def load_env_local():
    env_path = os.path.join(ROOT_DIR, ".env.local")
    if not os.path.exists(env_path):
        return
    with open(env_path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                key, _, val = line.partition("=")
                os.environ.setdefault(key.strip(), val.strip())


def parse_str_date(s, prev_date_str):
    """文字列形式の '5月17日' / ' 8月19日' を yyyy-mm-dd に変換する。
    年は直前のレコードの年を初期値として、月が大きく逆行したら+1年。"""
    if not isinstance(s, str):
        return None
    m = re.search(r"(\d{1,2})月\s*(\d{1,2})日", s.strip())
    if not m:
        return None
    mon = int(m.group(1))
    day = int(m.group(2))
    if not (1 <= mon <= 12 and 1 <= day <= 31):
        return None
    # 年推定: 直前の日付が無ければ2014年（このExcelは2014年から）、あれば直前の年を継承
    if prev_date_str:
        prev_year, prev_mon, _ = map(int, prev_date_str.split("-"))
        # 月が大きく前進したら同年、月が小さくなったら翌年
        year = prev_year + (1 if mon < prev_mon - 6 else 0)
    else:
        year = 2014
    return f"{year:04d}-{mon:02d}-{day:02d}"


def extract_records(path):
    try:
        import openpyxl
    except ImportError:
        print("openpyxl が未インストールです: pip install openpyxl")
        sys.exit(1)

    print(f"Excelを読み込み中: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    records = []
    current = {}
    prev_date_str = None
    skipped_count = 0  # 全国アウト行で日付parseに失敗してスキップした件数

    for row in rows:
        label = row[1]
        value = row[2]
        serial = row[3]

        if label == "全国アウト":
            date_str = None
            if isinstance(serial, int):
                date_str = serial_to_date(serial)
            elif hasattr(serial, "strftime"):
                date_str = serial.strftime("%Y-%m-%d")
            elif isinstance(serial, str):
                date_str = parse_str_date(serial, prev_date_str)
            if not date_str:
                skipped_count += 1
                current = {}
                continue
            prev_date_str = date_str
            current = {
                "date": date_str,
                "avg_in": float(value) if isinstance(value, (int, float)) else None,
            }
        elif label == "全国売上" and current:
            current["national_sales"] = float(value) if isinstance(value, (int, float)) else None
        elif label == "粗利" and current:
            current["gross_profit"] = float(value) if isinstance(value, (int, float)) else None
        elif label == "単価" and current:
            current["coin_price"] = float(value) if isinstance(value, (int, float)) else None
        elif label == "玉粗利" and current:
            current["coin_profit"] = float(value) if isinstance(value, (int, float)) else None
            # payout_rate は import_sis.py が機種別IN加重平均で書き込むので、ここでは触らない
            records.append(current)
            current = {}

    # SIS側の欠測日を検出して数を出す（値は書き換えない＝原典のまま入れる）。
    # 全国平均アウトが3,000未満の日は物理的にありえない水準で、原典Excelにも同じ値がある。
    # 集計側（稼働値の分母）では除外しているので、ここでは件数が急に増えたら気づけるようにする。
    low = [r for r in records if r.get("avg_in") is not None and r["avg_in"] < 3000]
    if low:
        print(f"⚠ 全国平均アウトが3,000未満の日: {len(low)}件（SIS側の欠測。集計では除外される）")
        for r in low[-5:]:
            print(f"    {r['date']} avg_in={r['avg_in']:.0f}")

    # 健全性チェック: パース失敗・全体件数・最終日付の警告を出す
    if skipped_count > 0:
        print(f"⚠️ 日付parseに失敗してスキップした行: {skipped_count} 件（仕様変更の可能性あり）", file=sys.stderr)
    if not records:
        print("❌ 抽出件数0件。日毎稼働全体.xlsx のシート構造を確認してください。", file=sys.stderr)
    else:
        latest = max(r["date"] for r in records)
        today = datetime.now().strftime("%Y-%m-%d")
        days_behind = (datetime.strptime(today, "%Y-%m-%d") - datetime.strptime(latest, "%Y-%m-%d")).days
        if days_behind > 3:
            print(f"⚠️ 最新データが {days_behind}日前 ({latest})。Excelの更新を確認してください。", file=sys.stderr)
        # null率チェック（payout_rateは import_sis.py 側で書き込まれるため、ここではチェックしない）
        null_sales = sum(1 for r in records if r.get("national_sales") is None)
        if null_sales / len(records) > 0.5:
            print(f"⚠️ national_salesがnullのレコードが {null_sales}/{len(records)} 件。抽出ロジック要確認。", file=sys.stderr)

    return records


def upsert_records(records, api_key):
    headers = {
        "apikey": api_key,
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates",
    }
    chunk_size = 500
    total = 0
    for i in range(0, len(records), chunk_size):
        chunk = records[i:i + chunk_size]
        res = requests.post(
            f"{SUPABASE_URL}/rest/v1/sis_national_daily",
            headers=headers,
            data=json.dumps(chunk),
        )
        if res.status_code in (200, 201):
            total += len(chunk)
            print(f"  アップロード {total}/{len(records)} 件")
        else:
            print(f"  エラー ({res.status_code}): {res.text[:200]}")
    print(f"完了: {total} 件")


def extract_records_from_weekly(path):
    try:
        import openpyxl
    except ImportError:
        print("openpyxl が未インストールです: pip install openpyxl")
        sys.exit(1)

    if not os.path.exists(path.replace("/", "\\")):
        print(f"スキップ: {path} が見つかりません")
        return []

    print(f"週次ファイルから全国アウトを読み込み中: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=True)
    sheets = [s for s in wb.sheetnames if s[:2].isdigit()]
    records = []
    for sheet_name in sheets:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4:
            continue
        dates = rows[2][6:13]
        nat_out = rows[3][6:13]
        for d, v in zip(dates, nat_out):
            if d and isinstance(v, (int, float)):
                records.append({"date": d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else str(d)[:10], "avg_in": float(v)})
    wb.close()
    return records


def run():
    load_env_local()

    api_key = os.environ.get("SUPABASE_SERVICE_KEY") or os.environ.get("VITE_SUPABASE_ANON_KEY")
    if not api_key:
        print("エラー: VITE_SUPABASE_ANON_KEY が未設定です")
        sys.exit(1)

    win_path = EXCEL_PATH.replace("/", "\\")
    if not os.path.exists(win_path):
        print(f"エラー: ファイルが見つかりません: {EXCEL_PATH}")
        sys.exit(1)

    records = extract_records(EXCEL_PATH)
    weekly_records = extract_records_from_weekly(WEEKLY_PATH)

    # 週次ファイルのデータで上書きマージ（日付をキーに）
    merged = {r["date"]: r for r in records}
    for r in weekly_records:
        merged[r["date"]] = {**merged.get(r["date"], {}), **r}
    # payout_rate は **絶対にこのリストへ入れない**。このスクリプトは原典に全国出玉率の行が無いので
    # 値を持っておらず、キーを含めると upsert が payout_rate=None を送り、同じBAT内で先に走った
    # import_sis.py が書いた値(機種別IN加重平均)を毎回消してしまう(2026-08-18まで実際に消えていた)。
    all_keys = {"date", "avg_in", "national_sales", "gross_profit", "coin_price", "coin_profit"}
    records = sorted([{k: v.get(k) for k in all_keys} for v in merged.values()], key=lambda r: r["date"])

    print(f"\n抽出合計: {len(records)} 件")
    if records:
        print(f"期間: {records[0]['date']} 〜 {records[-1]['date']}")
        print(f"サンプル: {records[-1]}")

    if not records:
        print("レコードなし。終了します。")
        return

    print("\nSupabase にアップロード中...")
    upsert_records(records, api_key)


if __name__ == "__main__":
    run()
