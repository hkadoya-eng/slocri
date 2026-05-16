"""
スマスロ モンスターストライク（仮）企画プロット Excel生成
対象読者: モンストをほぼ知らない人 × パチスロはそれなりに詳しい人
"""
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import points_to_pixels

OUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
           "proposals", "スマスロ_モンスト企画プロット.xlsx")
os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)

wb = Workbook()

# ── スタイル定義 ───────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=10, color="1A1A1A", name="メイリオ"):
    return Font(bold=bold, size=size, color=color, name=name)

def border(style="thin", color="CCCCCC"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

C_HEADER_BG  = "1B3A6B"   # 濃紺
C_HEADER_TXT = "FFFFFF"
C_MONST_ORG  = "E8620A"   # モンストオレンジ
C_MONST_LITE = "FDE8D8"   # 薄オレンジ
C_SECTION_BG = "2C6FAC"   # 中青
C_SECTION_TXT= "FFFFFF"
C_ROW_A      = "EBF4FF"   # 薄青（偶数行）
C_ROW_B      = "FFFFFF"   # 白（奇数行）
C_LABEL_BG   = "F0F4F8"   # ラベル背景
C_AT1        = "D5F5E3"   # 超絶（緑）
C_AT2        = "FEF9E7"   # 爆絶（黄）
C_AT3        = "FDEDEC"   # 轟絶（赤）
C_END        = "EAD6F8"   # エンディング（紫）
C_NOTE       = "FFF9E6"   # メモ背景（薄黄）
C_MONST_WORD = "E8620A"   # モンスト用語ハイライト色（テキスト）

thin_b = border("thin", "BBBBBB")
med_b  = border("medium", "888888")

def set_col_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_cell(ws, row, col, value, bg=None, txt_color="1A1A1A",
               bold=False, size=10, h_align="left", wrap=True, merge_end=None,
               border_style=None, italic=False):
    cell = ws.cell(row=row, column=col, value=value)
    if bg:
        cell.fill = fill(bg)
    cell.font = Font(bold=bold, size=size, color=txt_color, name="メイリオ", italic=italic)
    cell.alignment = align(h_align, "center", wrap)
    cell.border = border_style or thin_b
    if merge_end:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=merge_end[0], end_column=merge_end[1])
    return cell

def section_header(ws, row, col_start, col_end, text, bg=C_SECTION_BG):
    write_cell(ws, row, col_start, text, bg=bg, txt_color=C_SECTION_TXT,
               bold=True, size=11, h_align="center",
               merge_end=(row, col_end), border_style=med_b)
    ws.row_dimensions[row].height = 22

def note_row(ws, row, col_start, col_end, text):
    write_cell(ws, row, col_start, f"📌 {text}", bg=C_NOTE, txt_color="7D6608",
               size=8.5, italic=True, merge_end=(row, col_end))
    ws.row_dimensions[row].height = 30


# ══════════════════════════════════════════════════════════════
#  Sheet 1: 企画概要
# ══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "①企画概要"
set_col_width(ws1, [18, 55, 30])

# タイトルバー
write_cell(ws1, 1, 1, "スマスロ モンスターストライク（仮）企画プロット",
           bg=C_HEADER_BG, txt_color=C_HEADER_TXT, bold=True, size=14,
           h_align="center", merge_end=(1, 3), border_style=med_b)
ws1.row_dimensions[1].height = 32

write_cell(ws1, 2, 1, "対象読者：モンストをほぼ知らない人 ／ パチスロはそれなりに詳しい人",
           bg="F0F4F8", txt_color="555555", size=8.5, italic=True,
           h_align="center", merge_end=(2, 3))
ws1.row_dimensions[2].height = 18

row = 4
section_header(ws1, row, 1, 3, "■ IPについて（モンスト知らない人向け）"); row += 1
items = [
    ("モンスターストライク とは",
     "スマホゲームの大ヒット作（2013年〜）。最大4人でモンスターを弾いて敵を倒す協力ゲーム。"
     "ピーク時の2016年は月間アクティブユーザー5,000万人超。\n"
     "略称「モンスト」。「モン廃」と呼ばれる熱狂的プレイヤーを大量に生み出した。",
     ""),
    ("オーブ",
     "ゲーム内通貨。5個使うと1回ガチャ（抽選）ができる。50個で10連ガチャ。\n"
     "プレイヤーにとってオーブ＝最も大切なリソース。「オーブを貯めてガチャを引く」がプレイサイクルの核心。",
     "★ 本機の図柄にオーブをデザイン"),
    ("ガチャ",
     "ランダムでキャラクターを入手する抽選システム。★3〜★6の6段階レア度あり。\n"
     "★6は最高レア。「金オーブ」演出が出ると★6確定。モン廃の最大興奮ポイント。",
     "★ 本機でキャラ獲得＝仲間加入として設計"),
    ("マルチプレイ",
     "最大4人でクエストに挑む協力プレイ。「仲間を集めてクエストに挑む」がモンストの体験核心。",
     "★ 4人集めてATスタートの設計根拠"),
    ("超絶・爆絶・轟絶",
     "モンスト内の高難度クエストの格付け。超絶＜爆絶＜轟絶の順で難しく報酬も大きい。\n"
     "当時クリアできなかった超絶攻略に再挑戦する感覚を本機のATに再現する。",
     "★ AT格付けに直接使用"),
    ("覇者の塔",
     "月次で開催される高難度コンテンツ。フロアを登っていく形式で段階的に報酬が増える。\n"
     "達成感が強く、モン廃層に最も記憶に残るコンテンツのひとつ。",
     "★ AT進行の視覚的指標に使用"),
    ("映画（2作品）",
     "①はじまりの場所へ（2016年）主人公レンが父を探す旅。ゲノムがラスボス。\n"
     "②ソラノカナタ（2018年）母マナとの13年ぶりの再会。センジュがラスボス。\n"
     "親子の絆・仲間の友情がテーマの感動作。",
     "★ 映画シーンをATのストーリー演出に使用"),
]
for label, body, note in items:
    bg = C_ROW_A if row % 2 == 0 else C_ROW_B
    write_cell(ws1, row, 1, label, bg=C_LABEL_BG, bold=True, size=9.5)
    write_cell(ws1, row, 2, body, bg=bg, size=9)
    write_cell(ws1, row, 3, note, bg=C_MONST_LITE, txt_color=C_MONST_ORG,
               bold=True, size=8.5, h_align="center")
    ws1.row_dimensions[row].height = 52
    row += 1

row += 1
section_header(ws1, row, 1, 3, "■ コンセプト・ターゲット"); row += 1
concept_items = [
    ("機種名（仮）", "スマスロ モンスターストライク", ""),
    ("キャッチコピー",
     "「仲間を集めて、クエストへ。あの頃の熱狂が、リールに宿る。」", ""),
    ("コアコンセプト",
     "モンストの「オーブを貯めてガチャを引き、4人の仲間を集めてクエストに挑む」\n"
     "というプレイサイクルをそのままパチスロのゲーム性に落とし込む。\n"
     "映画のストーリーが覇者の塔（AT）の進行とリンクし、感動と出玉が同時に訪れる設計。", ""),
    ("主ターゲット",
     "超絶攻略世代（元モン廃）：現在30〜35歳。2015〜17年頃にモンストに熱中。\n"
     "今はパチスロをたまに打つ世代。「あいつも誘って打ちに行こう」が来店動機になる層。", ""),
    ("副ターゲット",
     "スマスロ世代（25〜28歳）：モンストは名前くらい知っている。\n"
     "IP依存より新しいゲーム性で選ぶ。オーブ収集→ガチャの設計が刺さる。", ""),
    ("ベンチマーク",
     "スマスロ まどか外伝（ストーリーAT・段階昇格設計）\n"
     "スマスロ 北斗の拳（CZ→AT→上位ATの王道・高射幸性）\n"
     "スマスロ からくりサーカス（段階的昇格・セット継続設計）", ""),
    ("射幸性の所在",
     "層①（入口）バウンドが止まらない緊張感（CZ演出）\n"
     "層②（中核）爆絶・轟絶への昇格決定の瞬間（メイン射幸性）\n"
     "層③（長期）映画完走・エンディング到達（夢の体験）", ""),
]
for label, body, note in concept_items:
    bg = C_ROW_A if row % 2 == 0 else C_ROW_B
    write_cell(ws1, row, 1, label, bg=C_LABEL_BG, bold=True, size=9.5)
    write_cell(ws1, row, 2, body, bg=bg, size=9)
    write_cell(ws1, row, 3, note, bg=bg, size=9)
    ws1.row_dimensions[row].height = 52
    row += 1


# ══════════════════════════════════════════════════════════════
#  Sheet 2: ゲームフロー
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("②ゲームフロー")
set_col_width(ws2, [20, 38, 38, 14])

write_cell(ws2, 1, 1, "ゲームフロー全体図",
           bg=C_HEADER_BG, txt_color=C_HEADER_TXT, bold=True, size=13,
           h_align="center", merge_end=(1, 4), border_style=med_b)
ws2.row_dimensions[1].height = 28

# ヘッダー行
for ci, h in enumerate(["フェーズ", "プレイヤーが見ること・体験すること",
                         "設計の意図（モンスト要素の翻訳）", "モンスト対応"], 1):
    write_cell(ws2, 2, ci, h, bg=C_SECTION_BG, txt_color=C_SECTION_TXT,
               bold=True, size=10, h_align="center")
ws2.row_dimensions[2].height = 20

flow = [
    ("通常時\n（クエスト進行中）",
     "毎ゲーム、リールにオーブ図柄が狙える。\n"
     "弱点属性のオーブ図柄が揃うと大ダメージ。\n"
     "ボスのHPゲージが画面に表示され、じわじわ減っていく。",
     "通常時＝クエスト進行中という概念に統一。\n"
     "「毎ゲームがゲームを進めている」感覚を作る。\n"
     "退屈な待ち時間を消す設計。",
     "クエスト進行\n属性/弱点",
     C_ROW_A),
    ("オーブ収集\n→ガチャ発動",
     "オーブ図柄が3つ揃うとオーブカウンターに蓄積。\n"
     "5個溜まったらガチャ演出が発生。\n"
     "仮面のキャラがくるくる回転して止まる演出。\n"
     "★5か★6かで歓声が変わる。",
     "モンストの「オーブ5個＝1ガチャ」ルールをそのまま採用。\n"
     "モン廃層には体に染みついた感覚。\n"
     "「あと何個」が常に意識される設計。",
     "オーブ\n単発ガチャ",
     C_ROW_B),
    ("仲間が集まる\n（4人揃うまで）",
     "ガチャを引くたびにキャラクターが1人加入。\n"
     "画面下部にパーティ枠（4つ）が表示され埋まっていく。\n"
     "3人目まで埋まった「あと1人」状態が最大の緊張感。",
     "モンストのマルチは最大4人。\n"
     "「4人揃えてクエストへ」という体験をそのまま再現。\n"
     "仲間の顔ぶれでCZの格が決まる。",
     "4人マルチ\nパーティ編成",
     C_ROW_A),
    ("10連ガチャ\n（特別ルート）",
     "虹オーブ・ゾーン・天井で突如発動。\n"
     "10体のキャラが一斉に並ぶ豪華演出。\n"
     "その中からランク上位4人が自動でピックアップされる。\n"
     "通常より必ず良いパーティが組まれ上位CZ直行。",
     "10連ガチャはモン廃の最大興奮体験。\n"
     "自動ピックアップにすることで規制上もクリア。\n"
     "特別ルートとして希少性を担保。",
     "10連ガチャ\n限定演出",
     C_MONST_LITE),
    ("CZ\n（ボス戦：バウンド型）",
     "弾（モンスト玉）を弾く演出が始まる。\n"
     "弾がバウンドするたびにボスへのダメージが蓄積。\n"
     "「まだ跳ねるか？」の緊張感。\n"
     "多段バウンドでボス撃破→AT突入確定。",
     "モンストの「弾く→跳ね返る」コア体験を再現。\n"
     "既存9CZパターンにない「バウンド数が期待度を決める」新設計。\n"
     "パーティが強いほどバウンド上限が増える連動設計。",
     "ストライクショット\nバウンド体験",
     C_ROW_B),
    ("AT\n（覇者の塔：超絶）\n1〜10F",
     "映画①「はじまりの場所へ」の前半シーンが進行。\n"
     "覇者の塔のフロアを登りながらストーリーが展開。\n"
     "5Fで爆絶昇格チャンス、10Fで轟絶昇格。\n"
     "感動シーンで純増が最大化する演出。",
     "「塔を登る達成感」×「映画の続きが見たい」の複合設計。\n"
     "出玉と感動が同時に訪れる瞬間を設計の核心に置く。",
     "覇者の塔\n超絶クエスト",
     C_AT1),
    ("AT昇格\n（爆絶）\n11〜30F",
     "映画①後半→映画②前半のシーンへ。\n"
     "父との再会シーン（18F）で友情コンボ大連鎖、純増MAX。\n"
     "30Fで轟絶昇格分岐。",
     "感動の山（18F父の再会）と緊張の谷（15Fゲノム激闘）を\n"
     "交互に設計し、感情の起伏がそのまま出玉変動になる。",
     "爆絶クエスト\n映画①",
     C_AT2),
    ("AT最上位\n（轟絶）\n31〜50F",
     "映画②クライマックス。センジュとの最終決戦。\n"
     "38FでカナタとマナのSS発動、純増爆増。\n"
     "50Fで映画ラストシーン。",
     "最大感動シーンに最大出玉を重ねる設計。\n"
     "「感動しながら出玉が伸びる」体験が他機種にない価値。",
     "轟絶クエスト\n映画②",
     C_AT3),
    ("エンディング\n（完走）",
     "両映画の全キャラが集合するエンドロール演出。\n"
     "覇者の塔全制覇の称号が付与される。\n"
     "「また来たい」余韻を残すラストシーン。",
     "設定6専用の夢体験。\n"
     "「エンディングを見た」がリピート来店・口コミの動機になる。",
     "覇者の塔\n全制覇",
     C_END),
]

row = 3
for phase, experience, intent, monst, bg in flow:
    write_cell(ws2, row, 1, phase, bg=C_LABEL_BG, bold=True, size=9.5, h_align="center")
    write_cell(ws2, row, 2, experience, bg=bg, size=9)
    write_cell(ws2, row, 3, intent, bg=bg, size=9)
    write_cell(ws2, row, 4, monst, bg=C_MONST_LITE, txt_color=C_MONST_ORG,
               bold=True, size=8.5, h_align="center")
    ws2.row_dimensions[row].height = 68
    row += 1


# ══════════════════════════════════════════════════════════════
#  Sheet 3: オーブ・ガチャ設計
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("③オーブ・ガチャ設計")
set_col_width(ws3, [20, 28, 28, 22])

write_cell(ws3, 1, 1, "オーブ・ガチャ・仲間集め 設計詳細",
           bg=C_HEADER_BG, txt_color=C_HEADER_TXT, bold=True, size=13,
           h_align="center", merge_end=(1, 4), border_style=med_b)
ws3.row_dimensions[1].height = 28

row = 3
section_header(ws3, row, 1, 4, "■ オーブの種類と価値"); row += 1
note_row(ws3, row, 1, 4,
         "モンスト本家のオーブ：5個で単発ガチャ / 50個で10連ガチャ。"
         "本機ではオーブを図柄としてデザインし、パチスロの役と連動させる。"); row += 1

for ci, h in enumerate(["オーブ種類", "入手契機（パチスロ）", "効果", "モンスト的価値"], 1):
    write_cell(ws3, row, ci, h, bg=C_SECTION_BG, txt_color=C_SECTION_TXT,
               bold=True, h_align="center", size=9.5)
ws3.row_dimensions[row].height = 18; row += 1

orb_data = [
    ("🔵 青オーブ（通常）", "3図柄揃い（通常役）", "オーブカウンター +1",
     "最も基本的なオーブ。コツコツ集める"),
    ("🔴 属性オーブ（色付き）", "弱点属性の3図柄揃い", "オーブカウンター +2",
     "弱点を突いた！という感覚を再現"),
    ("🌈 虹オーブ（最高）", "レア役・ゾーン到達・天井", "カウンター +5（即ガチャ）\nor 10連ガチャ発動",
     "モン廃が最も喜ぶ演出。金オーブ的な存在"),
]
for od in orb_data:
    bg = C_ROW_A if row % 2 == 0 else C_ROW_B
    for ci, v in enumerate(od, 1):
        write_cell(ws3, row, ci, v, bg=bg, size=9)
    ws3.row_dimensions[row].height = 38
    row += 1

row += 1
section_header(ws3, row, 1, 4, "■ 2つのガチャルート比較"); row += 1

for ci, h in enumerate(["ルート", "発動条件", "ガチャ内容", "期待できるAT格"], 1):
    write_cell(ws3, row, ci, h, bg=C_SECTION_BG, txt_color=C_SECTION_TXT,
               bold=True, h_align="center", size=9.5)
ws3.row_dimensions[row].height = 18; row += 1

gacha_data = [
    ("通常ルート\n（オーブ積み上げ）",
     "5オーブ毎に単発ガチャ発動\n4回引くと4人揃いCZへ",
     "1体ずつキャラが出現\n★5〜★6がランダムで登場\nキャラの顔ぶれでCZ格が変化",
     "★5混じり → 超絶AT\n★6揃い  → 爆絶AT\n限定キャラ → 轟絶確定",
     C_AT1),
    ("特別ルート\n（10連ガチャ）",
     "虹オーブ / ゾーン到達 / 天井\n突如発動（プレイヤーは選べない）",
     "10体のキャラが一斉に並ぶ豪華演出\n自動でランク上位4人をピックアップ\n通常より必ず良いパーティ構成に",
     "爆絶AT以上が確定\n全員★6 → 轟絶AT直行\n限定キャラ含む → エンディング射程",
     C_AT2),
]
for label, cond, content, expect, bg in gacha_data:
    write_cell(ws3, row, 1, label, bg=C_LABEL_BG, bold=True, size=9.5, h_align="center")
    write_cell(ws3, row, 2, cond, bg=bg, size=9)
    write_cell(ws3, row, 3, content, bg=bg, size=9)
    write_cell(ws3, row, 4, expect, bg=bg, size=9, bold=True)
    ws3.row_dimensions[row].height = 65
    row += 1

row += 1
section_header(ws3, row, 1, 4, "■ パーティ構成とCZ格の対応"); row += 1
note_row(ws3, row, 1, 4,
         "4人の顔ぶれ（レア度）がCZの難易度と直結し、攻略できればAT格が決まる。"); row += 1

for ci, h in enumerate(["パーティ構成", "CZ格", "突入AT", "AT期待枚数イメージ"], 1):
    write_cell(ws3, row, ci, h, bg=C_SECTION_BG, txt_color=C_SECTION_TXT,
               bold=True, h_align="center", size=9.5)
ws3.row_dimensions[row].height = 18; row += 1

party_data = [
    ("★5以下が1人以上いる", "通常CZ", "超絶AT", "150〜250枚", C_AT1),
    ("★6が2〜3人", "上位CZ", "爆絶AT", "400〜700枚", C_AT2),
    ("★6が4人揃う", "最上位CZ", "轟絶AT確定", "800〜1200枚", C_AT3),
    ("限定キャラ含む", "特別CZ", "轟絶 + エンディング射程", "1500枚超", C_END),
]
for comp, cz, at, exp, bg in party_data:
    write_cell(ws3, row, 1, comp, bg=C_LABEL_BG, bold=True, size=9)
    write_cell(ws3, row, 2, cz, bg=bg, size=9, h_align="center")
    write_cell(ws3, row, 3, at, bg=bg, size=9, bold=True, h_align="center")
    write_cell(ws3, row, 4, exp, bg=bg, size=9, h_align="center")
    ws3.row_dimensions[row].height = 28
    row += 1


# ══════════════════════════════════════════════════════════════
#  Sheet 4: AT設計（映画×覇者の塔）
# ══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("④AT設計（映画×覇者の塔）")
set_col_width(ws4, [12, 24, 30, 20, 14])

write_cell(ws4, 1, 1, "AT設計 ── 覇者の塔 × 映画ストーリー連動",
           bg=C_HEADER_BG, txt_color=C_HEADER_TXT, bold=True, size=13,
           h_align="center", merge_end=(1, 5), border_style=med_b)
ws4.row_dimensions[1].height = 28
note_row(ws4, 2, 1, 5,
         "覇者の塔（フロア制の高難度コンテンツ）を登りながら、映画のストーリーが進行する。"
         "感動シーンで純増が最大化する「感動と出玉の連動設計」が本機の最大の特徴。")
ws4.row_dimensions[2].height = 32

row = 4
for ci, h in enumerate(["フロア", "映画シーン", "感情・演出", "パチスロ演出", "AT格"], 1):
    write_cell(ws4, row, ci, h, bg=C_SECTION_BG, txt_color=C_SECTION_TXT,
               bold=True, h_align="center", size=9.5)
ws4.row_dimensions[row].height = 18; row += 1

at_scenes = [
    # 超絶AT
    ("1〜2F", "モンストグランプリ中にゲノム登場\n（映画①序章）",
     "懐かしさ・ワクワク\n▶ 掴みの興奮", "懐かしいSE・BGMで一気に引き込む", "超絶", C_AT1),
    ("3F", "時空を越えて4年前の世界へ\n謎の始まり",
     "驚き・謎への引き込み", "リール変化演出・画面転換", "超絶", C_AT1),
    ("5F", "謎のドラゴン・大人の陰謀\n【感情の谷①】",
     "緊張・不安\n▶ 継続するかのドキドキ",
     "継続演出で「続くか？」\n爆絶昇格チャンス発生", "超絶", C_AT1),
    ("7F", "仲間3人との絆が深まるシーン",
     "友情・高揚\n▶ 感動の布石",
     "友情コンボ初連鎖\n純増UP演出", "超絶", C_AT1),
    ("10F", "「はじまりの場所」を目指す決意",
     "高揚・期待\n▶ 第1章クリア",
     "爆絶昇格分岐演出\nチャプタークリアファンファーレ", "超絶→爆絶分岐", C_AT1),
    # 爆絶AT
    ("12F", "父・健太郎の痕跡を発見",
     "期待感・謎解きの高まり", "謎解き演出・小道具アイテム登場", "爆絶", C_AT2),
    ("15F", "ゲノムとの激闘・絶体絶命\n【感情の谷②】",
     "最大緊張・絶望感\n▶ 逆転への期待",
     "轟絶昇格チャンス演出\n「諦めるな」演出", "爆絶", C_AT2),
    ("18F", "★父・健太郎との再会シーン\n【感動の山①】",
     "感動・高揚・解放感\n▶ 本機最大の感情ピーク①",
     "友情コンボ大爆発\n純増MAX演出\nBGM最高潮", "爆絶", C_AT2),
    ("20F", "映画②へ転換\n旧東京・新東京の世界観",
     "新鮮な驚き・世界観の広がり",
     "画面・BGM・世界観が一変\nモンスト世界の深さを表現", "爆絶", C_AT2),
    ("25F", "カナタが母マナの存在を知る",
     "感動の布石・期待感の高まり",
     "BGMが静かに切り替わる\n次の感動への予告", "爆絶", C_AT2),
    ("30F", "センジュの正体が明かされる",
     "緊張・高揚・決戦への覚悟",
     "轟絶昇格分岐\nラスボス登場演出", "爆絶→轟絶分岐", C_AT2),
    # 轟絶AT
    ("33F", "センジュとの最終決戦開始",
     "最高の緊張感・決戦の高揚",
     "ボス戦BGM\nバウンド演出再び", "轟絶", C_AT3),
    ("35F", "旧東京落下の危機\n【感情の谷③】",
     "絶体絶命・諦めかけ\n▶ 逆転チャンス",
     "継続演出「最後の希望」\n逆転昇格の可能性", "轟絶", C_AT3),
    ("38F", "★カナタとマナ 13年ぶりの再会\n【感動の山②】",
     "最大の感動・涙・解放感\n▶ 本機最大の感情ピーク②",
     "SS発動演出\n純増爆増・画面エフェクト最大\n映画の名シーン完全再現", "轟絶", C_AT3),
    ("42F", "ソラが旧東京から飛び出す決死の行動",
     "息をのむ緊張・勇気への感動",
     "スロー演出\nリールが一時停止する特殊演出", "轟絶", C_AT3),
    ("45F", "センジュ撃破",
     "カタルシス・達成感・解放感",
     "全キャラSS連鎖演出\n爽快な勝利ファンファーレ", "轟絶", C_AT3),
    ("50F", "人間とモンスターが共に歩む新東京\n映画ラストシーン",
     "温かい感動・余韻\n▶ 完走への布石",
     "エンディング曲スタート\n全フロア制覇演出", "轟絶→ED分岐", C_AT3),
    # エンディング
    ("完走", "両映画の全キャラ集合エンドロール",
     "達成感・感謝・余韻\n▶ また来たいと思わせる",
     "覇者の塔全制覇称号\n「また会おう」メッセージ", "エンディング", C_END),
]

for floor, scene, emotion, prod, at_grade, bg in at_scenes:
    write_cell(ws4, row, 1, floor, bg=C_LABEL_BG, bold=True, size=9, h_align="center")
    write_cell(ws4, row, 2, scene, bg=bg, size=9)
    write_cell(ws4, row, 3, emotion, bg=bg, size=9)
    write_cell(ws4, row, 4, prod, bg=bg, size=9)
    write_cell(ws4, row, 5, at_grade, bg=bg, bold=True, size=9, h_align="center")
    ws4.row_dimensions[row].height = 45
    row += 1


# ══════════════════════════════════════════════════════════════
#  Sheet 5: スペック概要
# ══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("⑤スペック概要")
set_col_width(ws5, [22, 35, 30])

write_cell(ws5, 1, 1, "スペック概要（数値設計イメージ）",
           bg=C_HEADER_BG, txt_color=C_HEADER_TXT, bold=True, size=13,
           h_align="center", merge_end=(1, 3), border_style=med_b)
ws5.row_dimensions[1].height = 28
note_row(ws5, 2, 1, 3,
         "数値はあくまでイメージ。ミドル〜ハイミドルスペック想定。"
         "射幸性の中核は「初当りの重さ」ではなく「爆絶・轟絶への昇格」に置く。")
ws5.row_dimensions[2].height = 28

row = 4
section_header(ws5, row, 1, 3, "■ 基本スペック"); row += 1
spec_items = [
    ("CZ間初当り（超絶突入）", "1/240〜1/280前後（設定1〜6）", "ミドルスペック想定"),
    ("機械割", "97.5〜112.5%（設定1〜6）", "設定5で約107%"),
    ("超絶AT期待枚数", "200〜250枚", "通常完走ベース"),
    ("爆絶AT期待枚数", "500〜700枚", "中核の出玉帯"),
    ("轟絶AT期待枚数", "900〜1200枚", "上位体験"),
    ("エンディング期待枚数", "差枚数管理 1500〜2000枚超", "設定6の夢体験"),
    ("超絶→爆絶昇格率", "約35%", "ここが中核の射幸性"),
    ("爆絶→轟絶昇格率", "約15%", "希少体験として設計"),
    ("天井", "CZ間 最大999G（超絶確定）", "スイカ等の天井は別途検討"),
]
for label, value, note in spec_items:
    bg = C_ROW_A if row % 2 == 0 else C_ROW_B
    write_cell(ws5, row, 1, label, bg=C_LABEL_BG, bold=True, size=9.5)
    write_cell(ws5, row, 2, value, bg=bg, bold=True, size=10, h_align="center")
    write_cell(ws5, row, 3, note, bg=bg, size=9, txt_color="666666", italic=True)
    ws5.row_dimensions[row].height = 24
    row += 1

row += 1
section_header(ws5, row, 1, 3, "■ 射幸性の3層設計"); row += 1
layer_items = [
    ("層①（入口）",
     "バウンドが止まらない緊張感",
     "CZ中のバウンド演出。ライトユーザーでも楽しめる射幸性。"),
    ("層②（中核）★メイン",
     "爆絶・轟絶への昇格決定の瞬間",
     "ここに最大のエネルギーをかける。昇格演出が本機最大の興奮。"),
    ("層③（長期）",
     "映画完走・エンディング到達",
     "設定6専用の夢体験。口コミで広がる「あの台でエンディング見た」。"),
]
for layer, core, detail in layer_items:
    bg = C_ROW_A if row % 2 == 0 else C_ROW_B
    write_cell(ws5, row, 1, layer, bg=C_LABEL_BG, bold=True, size=9.5)
    write_cell(ws5, row, 2, core, bg=bg, bold=True, size=10)
    write_cell(ws5, row, 3, detail, bg=bg, size=9)
    ws5.row_dimensions[row].height = 36
    row += 1

row += 1
section_header(ws5, row, 1, 3, "■ ベンチマーク機種"); row += 1
bm_items = [
    ("スマスロ まどか外伝",
     "ストーリーAT・段階昇格設計のお手本",
     "感動と出玉の連動・穢れ→上位AT昇格の設計思想を参照"),
    ("スマスロ 北斗の拳",
     "CZ→AT→上位ATの王道設計・高射幸性",
     "CZ格付けと昇格確率の設計バランスを参照"),
    ("スマスロ からくりサーカス",
     "段階的昇格・ゲーム数固定CZ設計",
     "CZの長さとAT昇格演出のテンポ感を参照"),
]
for label, value, note in bm_items:
    bg = C_ROW_A if row % 2 == 0 else C_ROW_B
    write_cell(ws5, row, 1, label, bg=C_LABEL_BG, bold=True, size=9.5)
    write_cell(ws5, row, 2, value, bg=bg, bold=True, size=9.5)
    write_cell(ws5, row, 3, note, bg=bg, size=9)
    ws5.row_dimensions[row].height = 30
    row += 1


wb.save(OUT_PATH)
print(f"保存完了: {OUT_PATH}")
