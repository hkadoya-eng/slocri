"""
日毎稼働全体.xlsx から全国日次集計データを Supabase の sis_national_daily テーブルにインポートする。

使い方:
  python import_national_daily.py
"""

import sys
import io
import os
import json
import requests
from datetime import datetime, timedelta

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

EXCEL_PATH = "Z:/01_SISデータ/PS/日毎稼働全体.xlsx"
WEEKLY_PATH = "Z:/01_SISデータ/PS/PS日毎稼働まとめ_2026.xlsm"
SUPABASE_URL = "https://vpzbtuucopucablwyqeq.supabase.co"
SHEET_NAME = "他機種含む"

# Excelシリアル日付 → date文字列
EXCEL_EPOCH = datetime(1899, 12, 30)


def serial_to_date(serial):
    return (EXCEL_EPOCH + timedelta(days=int(serial))).strftime("%Y-%m-%d")


def load_env_local():
    env_path = os.path.join(os.path.dirname(__file__), ".env.local")
    if not os.path.exists(env_path):
        return
    with open(env_path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                key, _, val = line.partition("=")
                os.environ.setdefault(key.strip(), val.strip())


def extract_records(path):
    try:
        import openpyxl
    except ImportError:
        print("openpyxl が未インストールです: pip install openpyxl")
        sys.exit(1)

    print(f"Excelを読み込み中: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    records = []
    current = {}

    for row in rows:
        label = row[1]
        value = row[2]
        serial = row[3]

        if label == "全国アウト" and (isinstance(serial, int) or hasattr(serial, "strftime")):
            date_str = serial_to_date(serial) if isinstance(serial, int) else serial.strftime("%Y-%m-%d")
            current = {
                "date": date_str,
                "avg_in": float(value) if isinstance(value, (int, float)) else None,
            }
        elif label == "全国売上" and current:
            current["national_sales"] = float(value) if isinstance(value, (int, float)) else None
        elif label == "粗利" and current:
            current["gross_profit"] = float(value) if isinstance(value, (int, float)) else None
        elif label == "単価" and current:
            current["coin_price"] = float(value) if isinstance(value, (int, float)) else None
        elif label == "玉粗利" and current:
            current["coin_profit"] = float(value) if isinstance(value, (int, float)) else None
            # 出玉率を計算（暫定）
            avg_in = current.get("avg_in")
            cp = current.get("coin_price")
            gp = current.get("gross_profit")
            if avg_in and cp and gp is not None and avg_in * cp != 0:
                current["payout_rate"] = round((1 - gp / (avg_in * cp)) * 100, 2)
            else:
                current["payout_rate"] = None
            records.append(current)
            current = {}

    return records


def upsert_records(records, api_key):
    headers = {
        "apikey": api_key,
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates",
    }
    chunk_size = 500
    total = 0
    for i in range(0, len(records), chunk_size):
        chunk = records[i:i + chunk_size]
        res = requests.post(
            f"{SUPABASE_URL}/rest/v1/sis_national_daily",
            headers=headers,
            data=json.dumps(chunk),
        )
        if res.status_code in (200, 201):
            total += len(chunk)
            print(f"  アップロード {total}/{len(records)} 件")
        else:
            print(f"  エラー ({res.status_code}): {res.text[:200]}")
    print(f"完了: {total} 件")


def extract_records_from_weekly(path):
    try:
        import openpyxl
    except ImportError:
        print("openpyxl が未インストールです: pip install openpyxl")
        sys.exit(1)

    if not os.path.exists(path.replace("/", "\\")):
        print(f"スキップ: {path} が見つかりません")
        return []

    print(f"週次ファイルから全国アウトを読み込み中: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=True)
    sheets = [s for s in wb.sheetnames if s[:2].isdigit()]
    records = []
    for sheet_name in sheets:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4:
            continue
        dates = rows[2][6:13]
        nat_out = rows[3][6:13]
        for d, v in zip(dates, nat_out):
            if d and isinstance(v, (int, float)):
                records.append({"date": d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else str(d)[:10], "avg_in": float(v)})
    wb.close()
    return records


def run():
    load_env_local()

    api_key = os.environ.get("SUPABASE_SERVICE_KEY") or os.environ.get("VITE_SUPABASE_ANON_KEY")
    if not api_key:
        print("エラー: VITE_SUPABASE_ANON_KEY が未設定です")
        sys.exit(1)

    win_path = EXCEL_PATH.replace("/", "\\")
    if not os.path.exists(win_path):
        print(f"エラー: ファイルが見つかりません: {EXCEL_PATH}")
        sys.exit(1)

    records = extract_records(EXCEL_PATH)
    weekly_records = extract_records_from_weekly(WEEKLY_PATH)

    # 週次ファイルのデータで上書きマージ（日付をキーに）
    merged = {r["date"]: r for r in records}
    for r in weekly_records:
        merged[r["date"]] = {**merged.get(r["date"], {}), **r}
    all_keys = {"date", "avg_in", "national_sales", "gross_profit", "coin_price", "coin_profit", "payout_rate"}
    records = sorted([{k: v.get(k) for k in all_keys} for v in merged.values()], key=lambda r: r["date"])

    print(f"\n抽出合計: {len(records)} 件")
    if records:
        print(f"期間: {records[0]['date']} 〜 {records[-1]['date']}")
        print(f"サンプル: {records[-1]}")

    if not records:
        print("レコードなし。終了します。")
        return

    print("\nSupabase にアップロード中...")
    upsert_records(records, api_key)


if __name__ == "__main__":
    run()
