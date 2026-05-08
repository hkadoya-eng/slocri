"""
PS日毎稼働まとめ_2025/2026.xlsm から src/sisLibrary.json を再構築する。

使い方:
  python build_sis_library.py
"""

import sys, io, json, os
from datetime import datetime
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

PATHS = [
    "Z:/01_SISデータ/PS/PS日毎稼働まとめ_2025.xlsm",
    "Z:/01_SISデータ/PS/PS日毎稼働まとめ_2026.xlsm",
]
OUT_PATH = os.path.join(os.path.dirname(__file__), "src", "sisLibrary.json")


def extract_data(path):
    try:
        import openpyxl
    except ImportError:
        print("openpyxl未インストール")
        sys.exit(1)

    print(f"読み込み中: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=True)
    target_sheets = [s for s in wb.sheetnames if s[:2].isdigit()]
    print(f"  対象シート数: {len(target_sheets)}")

    # machine_name → {date_str: {out, coin, rate, profit, count}}
    records = defaultdict(dict)
    intros = {}

    for sheet_name in target_sheets:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 10:
            continue

        dates = rows[2][6:13]

        i = 4
        while i < len(rows) - 5:
            row = rows[i]
            machine = row[1]
            if machine and isinstance(machine, str) and machine.startswith("L") and "【" not in machine:
                machine = machine.strip()

                # 導入日取得（次行）
                intro_val = rows[i + 1][2] if len(rows) > i + 1 else None
                if isinstance(intro_val, datetime) and machine not in intros:
                    intros[machine] = intro_val.strftime("%Y-%m-%d")

                out_vals    = rows[i][6:13]
                coin_vals   = rows[i + 1][6:13]
                rate_vals   = rows[i + 2][6:13]
                profit_vals = rows[i + 3][6:13]
                count_vals  = rows[i + 5][6:13]

                for j, d in enumerate(dates):
                    if isinstance(d, datetime) and isinstance(out_vals[j], (int, float)):
                        ds = d.strftime("%Y-%m-%d")
                        records[machine][ds] = {
                            "out": int(out_vals[j]),
                            "coin": float(coin_vals[j]) if isinstance(coin_vals[j], (int, float)) else None,
                            "rate": float(rate_vals[j]) if isinstance(rate_vals[j], (int, float)) else None,
                            "profit": int(profit_vals[j]) if isinstance(profit_vals[j], (int, float)) else None,
                            "count": int(count_vals[j]) if isinstance(count_vals[j], (int, float)) else None,
                        }
                i += 6
            else:
                i += 1

    wb.close()
    return records, intros


def week_key(date_str):
    d = datetime.strptime(date_str, "%Y-%m-%d")
    # 週の月曜日に揃える
    mon = d - __import__("datetime").timedelta(days=d.weekday())
    return mon.strftime("%y%m%d")


def classify_decay(weekly_aout):
    if len(weekly_aout) < 4:
        return "データ不足"
    vals = [v for _, v in sorted(weekly_aout.items())]
    peak = max(vals)
    if peak == 0:
        return "不明"
    last = vals[-1]
    ratio = last / peak
    if ratio >= 0.85:
        return "安定稼働（W末でもピーク比85%以上）"
    elif ratio >= 0.6:
        return "中間減衰（W5で55-75%残存）"
    elif ratio >= 0.4:
        return "標準減衰（W8で40-60%残存）"
    else:
        return "急速減衰（W4以降で急落）"


def build_machine_entry(machine, day_data, intro):
    dates = sorted(day_data.keys())
    outs = [day_data[d]["out"] for d in dates if day_data[d]["out"]]
    coins = [day_data[d]["coin"] for d in dates if day_data[d]["coin"]]
    rates = [day_data[d]["rate"] for d in dates if day_data[d]["rate"]]
    profits = [day_data[d]["profit"] for d in dates if day_data[d]["profit"]]

    # 週次平均アウト
    weekly = defaultdict(list)
    for d in dates:
        if day_data[d]["out"]:
            weekly[week_key(d)].append(day_data[d]["out"])
    weekly_aout = {k: round(sum(v) / len(v), 1) for k, v in weekly.items()}

    return {
        "name": machine,
        "days": len(dates),
        "weeks": len(weekly_aout),
        "first_date": dates[0],
        "last_date": dates[-1],
        "intro": intro or dates[0],
        "avg_aout": round(sum(outs) / len(outs), 1) if outs else 0,
        "peak_aout": max(outs) if outs else 0,
        "avg_coin": round(sum(coins) / len(coins), 2) if coins else 0,
        "avg_rate": round(sum(rates) / len(rates), 1) if rates else 0,
        "avg_profit": round(sum(profits) / len(profits), 1) if profits else 0,
        "decay_pattern": classify_decay(weekly_aout),
        "weekly_aout": dict(sorted(weekly_aout.items())),
    }


def run():
    all_records = defaultdict(dict)
    all_intros = {}

    for path in PATHS:
        if not os.path.exists(path.replace("/", "\\")):
            print(f"スキップ（ファイルなし）: {path}")
            continue
        records, intros = extract_data(path)
        for m, days in records.items():
            all_records[m].update(days)
        for m, intro in intros.items():
            if m not in all_intros:
                all_intros[m] = intro

    print(f"\n機種数: {len(all_records)}")

    machines = {}
    for machine, day_data in sorted(all_records.items()):
        if len(day_data) < 3:
            continue
        machines[machine] = build_machine_entry(machine, day_data, all_intros.get(machine))

    # 期間を算出
    all_dates = []
    for m in machines.values():
        all_dates += [m["first_date"], m["last_date"]]
    period_start = min(all_dates)[:4] if all_dates else "2024"
    period_end = max(all_dates)[:4] if all_dates else "2026"

    library = {
        "source": "SIS実稼働データ",
        "period": f"{period_start}-{period_end}",
        "total": len(machines),
        "machines": machines,
    }

    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(library, f, ensure_ascii=False, indent=2)

    print(f"書き込み完了: {OUT_PATH}")
    print(f"機種数: {len(machines)} / 期間: {period_start}-{period_end}")
    dates = sorted(set(m["last_date"] for m in machines.values()))
    print(f"最新データ: {dates[-1] if dates else 'なし'}")


if __name__ == "__main__":
    run()
