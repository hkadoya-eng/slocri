"""
パチスロ CZ設計パターン一覧 Excel生成
"""
import os
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

OUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
           "proposals", "CZ設計パターン一覧.xlsx")
os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)

wb = Workbook()
ws = wb.active
ws.title = "CZ設計パターン一覧"

# ── カラー定義 ─────────────────────────────────────────────────
PATTERNS = [
    {
        "no": 1,
        "type": "ゲーム数固定型",
        "summary": "決まったG数内でAT抽選。\nシンプルで分かりやすく初心者にも直感的。",
        "color": "D6EAF8",  # 薄青
        "header": "1A5276",
        "machines": [
            ("スマスロ 北斗の拳", "夢想転生チャンス", "15G",
             "15G以内にAT（夢想転生）当選を目指す。期待度は設定や状態で変化。",
             "短期決戦の緊張感。当たるか外れるかがすぐ分かる爽快設計"),
            ("スマスロ からくりサーカス", "幕間チャンス", "10G+α",
             "10G固定。小役入賞で残りG数HOLD。2連続小役でAT当選。",
             "小役を引くたびに延命できる「自分で粘れる感覚」"),
            ("スマスロ 化物語", "解呪ノ儀", "モード別",
             "規定G数（100/200/300G）周期でCZ突入。レア小役でAT抽選。",
             "周期到達の達成感＋レア役の偶然性が交差する設計"),
        ]
    },
    {
        "no": 2,
        "type": "バトル型",
        "summary": "キャラ対戦でAT突入を争う。\nIP世界観と融合しやすく演出の幅が広い。",
        "color": "FDEDEC",  # 薄赤
        "header": "922B21",
        "machines": [
            ("スマスロ 東京リベンジャーズ", "リベンジチャンス", "可変",
             "全役でAT成功抽選。「リベンジ演出」発生で上位AT確定。対戦相手で期待度変化。",
             "IP愛着がある敵キャラへの「勝ちたい」感情が高揚感を生む"),
            ("スマスロ 炎炎ノ消防隊2", "炎炎激闘", "可変",
             "バトル勝利でボーナス獲得。連チャンの爽快感を重視した設計。",
             "戦闘演出の迫力と当否のドキドキが一体化"),
            ("スマスロ 東京喰種", "バトルCZ", "50G前後",
             "対戦相手で期待度が変化。上位AT（天国）確定の可能性あり。",
             "弱い相手＝ガッカリ、強い相手＝期待というIP知識が活きる設計"),
        ]
    },
    {
        "no": 3,
        "type": "セット継続型",
        "summary": "1セットごとに継続/終了を抽選。\n「もう1セット」の期待で引き付ける。",
        "color": "FEF9E7",  # 薄黄
        "header": "9A7D0A",
        "machines": [
            ("スマスロ 革命機ヴァルヴレイヴ2", "引き戻しゾーン", "66G×複数セット",
             "AT終了後の66G継続ゾーン。複数セット継続でAT復帰に期待。",
             "終わらない期待感。「まだ終わってない」の持続設計"),
            ("スマスロ 化物語", "解呪連モード", "セット型",
             "CZ当選時・AT終了後から移行。全役でCZ抽選しセットを重ねる。",
             "セットを積み上げるほど昂ぶる期待感"),
            ("スマスロ からくりサーカス", "懸糸傀儡", "3セット+",
             "真夜中のサーカス突入時は3セット継続濃厚。セット消化でAT確定。",
             "「3セット行ったら確定」という分かりやすいゴール設計"),
        ]
    },
    {
        "no": 4,
        "type": "自力型",
        "summary": "押し順・小役でAT抽選。\nプレイヤーに「自分で引いた」操作感を与える。",
        "color": "E8F8F5",  # 薄緑
        "header": "1A5632",
        "machines": [
            ("スマスロ 攻殻機動隊", "S.A.M.", "可変",
             "小役成立時に「回胴HACK」発生。ラフ目停止でエピソード進行。進行度で期待度変化。",
             "リールの出目を「読む」楽しさ。熟練者ほど興奮できる設計"),
            ("L真・一騎当千", "闘士決戦", "可変",
             "小役の組み合わせで敵を攻撃。攻撃パターン変化でAT突入率が変動。",
             "小役の引きが直接ダメージに変換される因果の明快さ"),
            ("Lアニマルスロットドッチ", "アニマルチャンス", "3回チャレンジ",
             "最大3回の押し順当てチャレンジ。全正解でST突入。",
             "押し順当ては「当たった！」の達成感が強い。ライトユーザー向け"),
        ]
    },
    {
        "no": 5,
        "type": "ループ型",
        "summary": "CZ自体がループしてAT突入率を高める。\n長期滞在でプレイヤーを引き付ける。",
        "color": "F4ECF7",  # 薄紫
        "header": "6C3483",
        "machines": [
            ("スマスロ ヨルムンガンド", "シューティングゾーン等3種", "可変",
             "3種類のCZがループ。「滅びの丘」突入でAT濃厚。AT終了後もCZでループに期待。",
             "CZを繰り返すうちに「いつか必ず当たる」という持続的期待を生む"),
            ("スマスロ 鉄拳6", "鉄拳チャンス", "15G×繰り返し",
             "15GのCZをボーナス後に繰り返し突入。拳奪バトルでボーナス継続。",
             "短期CZの繰り返しが連続当選感を演出"),
            ("スマスロ ありふれた職業で世界最強", "嫖王決定戦/香織復活チャンス", "周期型",
             "2種類のCZがモード別周期で交互に発生。上位モードでループ頻度UP。",
             "「次の周期では当たるかも」という継続モチベーション設計"),
        ]
    },
    {
        "no": 6,
        "type": "引き戻し型",
        "summary": "AT終了後の短いゾーンでAT復帰を狙う。\n「終わった」後の逆転体験が興奮を生む。",
        "color": "FDF2E9",  # 薄オレンジ
        "header": "A04000",
        "machines": [
            ("スマスロ 化物語", "夢の時間ヲ終わラセルな", "3G",
             "AT終了後3G。レア役でAT復帰、変則停止で大チャンス。超短期決戦。",
             "「たった3G」の極度の緊張感。終わった瞬間の逆転期待"),
            ("スマスロ 革命機ヴァルヴレイヴ2", "引き戻しゾーン", "66G",
             "AT終了後66G継続。紫ランプ点滅で突入示唆。",
             "長めの引き戻しゾーンで「まだチャンスがある」持続感"),
            ("スマスロ 攻殻機動隊", "白の境界（失敗後救済）", "天井短縮",
             "CZ失敗時に次回天井を400G+αに短縮。間接的引き戻し。",
             "「外れても損していない」というマイナスを緩和する設計"),
        ]
    },
    {
        "no": 7,
        "type": "昇格型",
        "summary": "CZ中に上位CZへ段階的に昇格する。\n「まだ上がある」という階層の深みが魅力。",
        "color": "EAFAF1",  # 薄ミント
        "header": "1E8449",
        "machines": [
            ("スマスロ 攻殻機動隊", "S.A.M.→白の境界", "エピソード進行",
             "S.A.M.（通常CZ）→エピソード昇格→「白の境界」（上位CZ）。レベル3まで段階UP。",
             "昇格のたびに期待度が上昇するRPG的な達成感"),
            ("スマスロ 鉄拳6", "CZモードA〜E", "5段階",
             "CZモードがA〜Eの5段階。レア役でモード昇格、モードが上がるほどAT期待度UP。",
             "モード昇格の瞬間が最大の興奮ポイント。段階の視覚化が重要"),
            ("スマスロ まどか☆マギカ外伝", "前兆→マギカラッシュ昇格", "段階型",
             "通常CZ→前兆→上位AT昇格。穢れシステムによる内部昇格も複合。",
             "外部から見えない内部昇格の存在が「実はもう昇格している」期待を生む"),
        ]
    },
    {
        "no": 8,
        "type": "ゾーン型",
        "summary": "特定G数帯でAT抽選が優遇される。\n立ち回りの奥行きを生みホール稼働に直結。",
        "color": "EBF5FB",  # 薄スカイ
        "header": "1A6A8A",
        "machines": [
            ("スマスロ ありふれた職業で世界最強", "300G周辺ゾーン", "末尾00G周期",
             "300G付近で高確率CZ突入。モードCなら期待度約80%。周期は内部モードで変化。",
             "ゾーン手前でのコイン投入判断が生まれる立ち回りの楽しさ"),
            ("スマスロ 東京喰種", "末尾00G・50Gゾーン", "50G周期",
             "内部モード6種類で周期が異なる。上位モードなら300G以内でもゾーン高期待度。",
             "「ここから打てば美味しい」という狙い目文化の醸成"),
            ("スマスロ 鉄拳6", "ゾーン+モード複合", "可変",
             "レア小役でモード昇格し通常ゾーン期待度が上昇。モード管理と連動。",
             "小役を引くたびにゾーンへの期待も高まる複合設計"),
        ]
    },
    {
        "no": 9,
        "type": "天井移行型",
        "summary": "規定G数到達で強制的にCZへ移行。\n「必ず来る」救済が長期プレイの安心感を生む。",
        "color": "F2F3F4",  # 薄グレー
        "header": "424949",
        "machines": [
            ("スマスロ ヨルムンガンド", "天井CZ", "1000G+α",
             "1000G+αで強制CZ移行。100G周期の高確移行でCZ突入チャンスも複数。",
             "天井の存在がリスク許容度を高め「打ち切れる安心感」を提供"),
            ("スマスロ 東京喰種", "CZ間天井・AT間天井", "600G/1200G",
             "CZ間天井600G（設定変更後200G短縮）、AT間天井1200G。2種類の天井が併存。",
             "天井の種類が多いほど狙い目が増えホール稼働向上に寄与"),
            ("スマスロ ありふれた職業で世界最強", "モード別天井", "モード依存",
             "モード別に規定G数が異なり天井到達でCZ確定。モード昇格で天井短縮。",
             "天井が動的に変化することで「今どこにいるか」の推測ゲームが生まれる"),
        ]
    },
]

# ── ヘッダー行 ─────────────────────────────────────────────────
HEADERS = [
    "No", "パターン分類", "設計の特徴・思想",
    "代表機種", "CZ名称", "ゲーム数/セット数",
    "仕組み・ゲーム性", "プレイヤー体験・設計価値"
]
COL_WIDTHS = [4, 16, 32, 24, 22, 14, 44, 40]

thin = Side(style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

header_font = Font(name="メイリオ", bold=True, color="FFFFFF", size=10)
header_fill = PatternFill("solid", fgColor="1C3A57")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ヘッダー
for ci, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border
    ws.column_dimensions[get_column_letter(ci)].width = w

ws.row_dimensions[1].height = 24

# ── データ行 ────────────────────────────────────────────────────
row = 2
for pat in PATTERNS:
    fill_bg  = PatternFill("solid", fgColor=pat["color"])
    fill_hdr = PatternFill("solid", fgColor=pat["header"])
    type_font = Font(name="メイリオ", bold=True, color="FFFFFF", size=10)
    body_font = Font(name="メイリオ", size=9)
    bold_font = Font(name="メイリオ", bold=True, size=9)
    wrap = Alignment(vertical="top", wrap_text=True)
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

    start_row = row

    for mi, (machine, cz_name, g_count, mechanism, experience) in enumerate(pat["machines"]):
        # No
        c = ws.cell(row=row, column=1, value=pat["no"] if mi == 0 else "")
        c.font = type_font; c.fill = fill_hdr; c.alignment = center_wrap; c.border = border

        # パターン分類
        c = ws.cell(row=row, column=2, value=pat["type"] if mi == 0 else "")
        c.font = type_font; c.fill = fill_hdr; c.alignment = center_wrap; c.border = border

        # 設計特徴
        c = ws.cell(row=row, column=3, value=pat["summary"] if mi == 0 else "")
        c.font = body_font; c.fill = fill_bg; c.alignment = wrap; c.border = border

        # 代表機種
        c = ws.cell(row=row, column=4, value=machine)
        c.font = bold_font; c.fill = fill_bg; c.alignment = wrap; c.border = border

        # CZ名称
        c = ws.cell(row=row, column=5, value=cz_name)
        c.font = body_font; c.fill = fill_bg; c.alignment = wrap; c.border = border

        # ゲーム数
        c = ws.cell(row=row, column=6, value=g_count)
        c.font = body_font; c.fill = fill_bg
        c.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        c.border = border

        # 仕組み
        c = ws.cell(row=row, column=7, value=mechanism)
        c.font = body_font; c.fill = fill_bg; c.alignment = wrap; c.border = border

        # プレイヤー体験
        c = ws.cell(row=row, column=8, value=experience)
        c.font = body_font; c.fill = fill_bg; c.alignment = wrap; c.border = border

        ws.row_dimensions[row].height = 52
        row += 1

    # パターン分類・No・設計特徴をマージ
    if len(pat["machines"]) > 1:
        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=row - 1, end_column=1)
        ws.merge_cells(start_row=start_row, start_column=2,
                       end_row=row - 1, end_column=2)
        ws.merge_cells(start_row=start_row, start_column=3,
                       end_row=row - 1, end_column=3)

    # パターン間の区切り行
    for ci in range(1, 9):
        c = ws.cell(row=row, column=ci, value="")
        c.fill = PatternFill("solid", fgColor="E8E8E8")
        c.border = border
    ws.row_dimensions[row].height = 6
    row += 1

# ── フリーズ・フィルター ─────────────────────────────────────────
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:H{row - 1}"

# ── 凡例シート ───────────────────────────────────────────────────
ws2 = wb.create_sheet("パターン解説")
ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 60

legend_title_font = Font(name="メイリオ", bold=True, size=12, color="1C3A57")
legend_head_fill  = PatternFill("solid", fgColor="1C3A57")
legend_head_font  = Font(name="メイリオ", bold=True, color="FFFFFF", size=10)

ws2.cell(row=1, column=1, value="CZ設計パターン 解説一覧").font = legend_title_font
ws2.merge_cells("A1:B1")

LEGEND = [
    ("ゲーム数固定型", "G数が決まっており短期決戦。結果がすぐ出るのでライトユーザーにも分かりやすい。"),
    ("バトル型", "IP世界観と直結しやすい。対戦相手やキャラで期待度を視覚化できる設計の王道。"),
    ("セット継続型", "「もう1セット」の期待感で引き付ける。終わらない感覚がコイン持ちを演出する。"),
    ("自力型", "押し順・小役で操作感を付与。「自分で引いた」体験が満足度を高める。熟練者向け。"),
    ("ループ型", "CZがループすることで長期滞在を促進。「いつか当たる」という信念を維持させる。"),
    ("引き戻し型", "AT終了後の短期チャンスで逆転体験を提供。損失を「まだある」に変換する設計。"),
    ("昇格型", "段階的に上位CZへ昇格。RPG的な達成感があり、昇格の瞬間が最大の興奮ポイント。"),
    ("ゾーン型", "特定G数帯で高確率。立ち回りの奥行きを生み、ホール稼働向上にも直結する設計。"),
    ("天井移行型", "規定G数で強制CZ移行。「必ず来る」という安心感が長期プレイを支える土台設計。"),
]

for ri, (name, desc) in enumerate(LEGEND, 3):
    pat_info = next((p for p in PATTERNS if p["type"] == name), None)
    bg = pat_info["color"] if pat_info else "FFFFFF"
    hc = pat_info["header"] if pat_info else "333333"

    c1 = ws2.cell(row=ri, column=1, value=name)
    c1.font = Font(name="メイリオ", bold=True, color="FFFFFF", size=9)
    c1.fill = PatternFill("solid", fgColor=hc)
    c1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c1.border = border

    c2 = ws2.cell(row=ri, column=2, value=desc)
    c2.font = Font(name="メイリオ", size=9)
    c2.fill = PatternFill("solid", fgColor=bg)
    c2.alignment = Alignment(vertical="center", wrap_text=True)
    c2.border = border

    ws2.row_dimensions[ri].height = 36

wb.save(OUT_PATH)
print(f"保存完了: {OUT_PATH}")
